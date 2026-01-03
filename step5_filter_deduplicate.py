#!/usr/bin/env python3
"""
Step 5: Filter and Deduplicate Data
Removes NA rows and deduplicates SD rows with data cleaning.
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, List, Dict, Tuple
import argparse
import sys
import shutil
from collections import defaultdict

from common.validation import validate_step5_input, FileValidator
from common.exceptions import TSConverterError
from common.config import get_clean_basename

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataFilter:
    """
    Data Filter for Step 5
    
    Four-stage filtering process:
    1. Remove rows where column H is NA/empty/"-"
    2. Find SD duplicate groups based on columns B,C,D,E,F,I,J similarity
    3. Clear columns K,L,M for all SD rows
    4. Deduplicate SD rows (keep first, remove others, set column N)
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Columns for comparison in SD deduplication
        self.comparison_columns = ['B', 'C', 'D', 'E', 'F', 'I', 'J']
        self.start_row = 11  # Start from row 11 (after headers and article data)
    
    def is_na_value(self, cell_value) -> bool:
        """
        Check if cell value should be considered as NA/empty
        
        Args:
            cell_value: Cell value to check
            
        Returns:
            True if value is NA/empty/"-"
        """
        if cell_value is None:
            return True
        
        if isinstance(cell_value, str):
            cleaned = cell_value.strip().upper()
            return cleaned in ["", "NA", "-"]
        
        return False
    
    def get_row_values(self, worksheet, row: int, columns: List[str]) -> tuple:
        """
        Get values from specified columns in a row
        
        Args:
            worksheet: openpyxl worksheet object
            row: Row number (1-based)
            columns: List of column letters
            
        Returns:
            Tuple of values from specified columns
        """
        values = []
        for col_letter in columns:
            col_num = openpyxl.utils.column_index_from_string(col_letter)
            cell_value = worksheet.cell(row, col_num).value
            # Normalize value for comparison
            if cell_value is None:
                values.append("")
            elif isinstance(cell_value, str):
                values.append(cell_value.strip())
            else:
                values.append(str(cell_value))
        
        return tuple(values)
    
    def has_meaningful_data(self, comparison_values: tuple) -> bool:
        """
        Check if tuple has at least one non-empty value for meaningful deduplication
        
        Args:
            comparison_values: Tuple of comparison values from columns B,C,D,E,F,I,J
            
        Returns:
            True if at least one value is non-empty, False if all are empty/whitespace
        """
        for val in comparison_values:
            if val and isinstance(val, str) and val.strip():
                return True
            elif val and not isinstance(val, str):
                return True
        return False
    
    def remove_na_rows(self, worksheet) -> int:
        """
        Remove rows where column H is NA/empty/"-"
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Number of rows removed
        """
        logger.info("Step 5.1: Removing NA rows from column H")
        
        h_col_num = openpyxl.utils.column_index_from_string('H')
        rows_to_delete = []
        
        # Find all rows to delete (process from bottom to top to avoid index issues)
        for row in range(worksheet.max_row, self.start_row - 1, -1):
            h_value = worksheet.cell(row, h_col_num).value
            
            if self.is_na_value(h_value):
                rows_to_delete.append(row)
                logger.debug(f"Marking row {row} for deletion (H = '{h_value}')")
        
        # Delete rows
        for row in rows_to_delete:
            worksheet.delete_rows(row, 1)
            logger.debug(f"Deleted row {row}")
        
        removed_count = len(rows_to_delete)
        logger.info(f"Removed {removed_count} NA rows")
        return removed_count
    
    def find_sd_duplicates(self, worksheet) -> Dict[tuple, List[int]]:
        """
        Find SD duplicate groups based on columns B,C,D,E,F,I,J similarity
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Dictionary mapping value tuples to list of row numbers
        """
        logger.info("Step 5.2: Finding SD duplicate groups")
        
        h_col_num = openpyxl.utils.column_index_from_string('H')
        duplicate_groups = defaultdict(list)
        
        # Counters for logging
        total_sd_rows = 0
        meaningful_sd_rows = 0
        empty_sd_rows = 0
        
        # Find all SD rows and group by comparison columns
        for row in range(self.start_row, worksheet.max_row + 1):
            h_value = worksheet.cell(row, h_col_num).value
            
            if h_value and isinstance(h_value, str) and h_value.strip().upper() == "SD":
                total_sd_rows += 1
                # Get comparison values
                comparison_values = self.get_row_values(worksheet, row, self.comparison_columns)
                
                # Only group rows with meaningful data for deduplication
                if self.has_meaningful_data(comparison_values):
                    duplicate_groups[comparison_values].append(row)
                    meaningful_sd_rows += 1
                    logger.debug(f"SD row {row} (meaningful): {comparison_values}")
                else:
                    empty_sd_rows += 1
                    logger.debug(f"SD row {row} (all empty - skipping deduplication): {comparison_values}")
        
        # Filter to only actual duplicates (groups with > 1 row)
        actual_duplicates = {k: v for k, v in duplicate_groups.items() if len(v) > 1}
        
        logger.info(f"SD Row Analysis:")
        logger.info(f"  Total SD rows found: {total_sd_rows}")
        logger.info(f"  Meaningful SD rows (for deduplication): {meaningful_sd_rows}")
        logger.info(f"  Empty SD rows (preserved individually): {empty_sd_rows}")
        logger.info(f"Found {len(actual_duplicates)} SD duplicate groups from meaningful rows")
        
        for group_key, rows in actual_duplicates.items():
            logger.debug(f"Group {group_key}: rows {rows}")
        
        return actual_duplicates
    
    def determine_common_value(self, worksheet, rows: List[int], column: str) -> str:
        """
        Determine common value from a group of rows for specified column
        
        Args:
            worksheet: openpyxl worksheet object
            rows: List of row numbers
            column: Column letter
            
        Returns:
            Common value or "Yearly" as default
        """
        col_num = openpyxl.utils.column_index_from_string(column)
        values = []
        
        for row in rows:
            cell_value = worksheet.cell(row, col_num).value
            if cell_value and isinstance(cell_value, str):
                cleaned_value = cell_value.strip()
                if cleaned_value:
                    values.append(cleaned_value)
        
        # Find most common value
        if values:
            from collections import Counter
            counter = Counter(values)
            most_common = counter.most_common(1)[0][0]
            logger.debug(f"Common value for column {column}: '{most_common}' from {values}")
            return most_common
        
        logger.debug(f"No common value found for column {column}, using default 'Yearly'")
        return "Yearly"
    
    def clear_sd_row_data(self, worksheet) -> int:
        """
        Step 5.3: Clear columns K, L, M for all rows with 'SD' in column H
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Number of SD rows processed
        """
        logger.info("Step 5.3: Clearing K,L,M for all SD rows")
        
        h_col_num = openpyxl.utils.column_index_from_string('H')
        cleared_count = 0
        
        # Process all rows to find SD entries
        for row in range(self.start_row, worksheet.max_row + 1):
            h_value = worksheet.cell(row, h_col_num).value
            
            if h_value and isinstance(h_value, str) and h_value.strip().upper() == "SD":
                # Clear columns K, L, M for this SD row
                for col_letter in ['K', 'L', 'M']:
                    col_num = openpyxl.utils.column_index_from_string(col_letter)
                    worksheet.cell(row, col_num).value = None
                    logger.debug(f"Cleared {col_letter}{row}")
                
                cleared_count += 1
                logger.debug(f"Processed SD row {row}")
        
        logger.info(f"Cleared K,L,M for {cleared_count} SD rows")
        return cleared_count
    
    def deduplicate_sd_rows(self, worksheet) -> int:
        """
        Step 5.4: Deduplicate SD rows
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Number of rows removed
        """
        duplicate_groups = self.find_sd_duplicates(worksheet)
        
        if not duplicate_groups:
            logger.info("No SD duplicates found")
            return 0
        
        rows_to_delete = []
        rows_processed = 0
        
        # Process each duplicate group
        for group_key, group_rows in duplicate_groups.items():
            logger.info(f"Processing duplicate group with {len(group_rows)} rows: {group_rows}")
            
            # Keep the first row, mark others for deletion
            keep_row = group_rows[0]
            delete_rows = group_rows[1:]
            rows_to_delete.extend(delete_rows)
            
            # Set column N to common value or "Yearly" for the kept row
            logger.debug(f"Keeping row {keep_row}, setting column N")
            n_col_num = openpyxl.utils.column_index_from_string('N')
            common_n_value = self.determine_common_value(worksheet, group_rows, 'N')
            worksheet.cell(keep_row, n_col_num).value = common_n_value
            logger.debug(f"Set N{keep_row} = '{common_n_value}'")
            
            rows_processed += 1
        
        # Delete duplicate rows (from bottom to top to avoid index issues)
        rows_to_delete.sort(reverse=True)
        for row in rows_to_delete:
            worksheet.delete_rows(row, 1)
            logger.debug(f"Deleted duplicate row {row}")
        
        removed_count = len(rows_to_delete)
        logger.info(f"Deduplicated {rows_processed} groups, removed {removed_count} duplicate rows")
        return removed_count
    
    def process_file(self, step4_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step4 file with filtering and deduplication
        
        Args:
            step4_file: Step4 input file path
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 DATAFILTER: Step 5 Filter and Deduplicate START")
        
        # DETAILED LOGGING: Initial state
        logger.info(f"🔧 DATAFILTER: DataFilter instance state:")
        logger.info(f"   - base_dir: {self.base_dir}")
        logger.info(f"   - output_dir: {self.output_dir}")
        logger.info(f"   - output_dir exists: {self.output_dir.exists()}")
        logger.info(f"   - current working directory: {Path.cwd()}")
        
        # Validate input file
        try:
            logger.info(f"📥 DATAFILTER: Validating input file: {step4_file}")
            validate_step5_input(step4_file)
            step4_path = Path(step4_file)
            logger.info(f"📥 DATAFILTER: Input validation passed")
            logger.info(f"   - step4_path: {step4_path}")
            logger.info(f"   - step4_path absolute: {step4_path.absolute()}")
            logger.info(f"   - step4_path exists: {step4_path.exists()}")
            logger.info(f"   - step4_path size: {step4_path.stat().st_size if step4_path.exists() else 'N/A'}")
        except TSConverterError as e:
            logger.error(f"❌ DATAFILTER: Input validation failed: {e}")
            raise
        
        # Enhanced output file handling with directory management
        if output_file is None:
            base_name = get_clean_basename(step4_path)
            output_file = self.output_dir / f"{base_name} - Step5.xlsx"
            logger.info(f"📤 DATAFILTER: Auto-generated output path: {output_file}")
        else:
            output_file = Path(output_file)
            logger.info(f"📤 DATAFILTER: Using provided output path: {output_file}")
            logger.info(f"   - output_file absolute: {output_file.absolute()}")
            logger.info(f"   - output_file parent: {output_file.parent}")
            logger.info(f"   - output_file parent exists: {output_file.parent.exists()}")
            
            # Ensure parent directory exists for provided output path
            if not output_file.parent.exists():
                logger.info(f"📂 DATAFILTER: Creating output directory: {output_file.parent}")
                output_file.parent.mkdir(parents=True, exist_ok=True)
                logger.info(f"📂 DATAFILTER: Output directory created: {output_file.parent.exists()}")
        
        # CRITICAL FIX: Always ensure output_file path is writable
        try:
            logger.info(f"🔒 DATAFILTER: Checking output path writability...")
            
            # For explicitly provided paths, create the directory structure if needed
            if output_file.parent != self.output_dir:
                logger.info(f"📂 DATAFILTER: Ensuring directory structure exists: {output_file.parent}")
                output_file.parent.mkdir(parents=True, exist_ok=True)
                logger.info(f"📂 DATAFILTER: Directory structure verified: {output_file.parent.exists()}")
            
            # Simple writability check
            if output_file.exists():
                logger.info(f"📄 DATAFILTER: Output file already exists, checking writability")
                if not os.access(output_file, os.W_OK):
                    raise TSConverterError(f"Output file is not writable: {output_file}")
                logger.info(f"📄 DATAFILTER: Existing output file is writable")
            
            if not os.access(output_file.parent, os.W_OK):
                raise TSConverterError(f"Output directory is not writable: {output_file.parent}")
                
            logger.info(f"✅ DATAFILTER: Output path validation successful: {output_file}")
        except TSConverterError as e:
            logger.error(f"❌ DATAFILTER: Output validation failed: {e}")
            raise
        
        logger.info(f"📝 DATAFILTER: Final file paths:")
        logger.info(f"   📥 Input: {step4_path}")
        logger.info(f"   📤 Output: {output_file}")
        
        # Copy Step4 file as starting point with enhanced error handling
        try:
            logger.info(f"📋 DATAFILTER: Copying Step4 file to output location...")
            logger.info(f"   - Source: {step4_path}")
            logger.info(f"   - Destination: {output_file}")
            logger.info(f"   - Source exists: {step4_path.exists()}")
            logger.info(f"   - Destination parent exists: {output_file.parent.exists()}")
            
            shutil.copy2(str(step4_path), str(output_file))
            
            logger.info(f"✅ DATAFILTER: File copy completed")
            logger.info(f"   - Output file exists: {output_file.exists()}")
            logger.info(f"   - Output file size: {output_file.stat().st_size if output_file.exists() else 'N/A'}")
            
        except (OSError, PermissionError) as e:
            logger.error(f"❌ DATAFILTER: Failed to copy input file to output location: {e}")
            logger.error(f"   - Source: {step4_path}")
            logger.error(f"   - Destination: {output_file}")
            logger.error(f"   - Source exists: {step4_path.exists()}")
            logger.error(f"   - Destination parent exists: {output_file.parent.exists()}")
            raise TSConverterError(f"Could not create output file: {str(e)}")
        
        # Load workbook with enhanced error handling
        try:
            logger.info(f"📊 DATAFILTER: Loading workbook from: {output_file}")
            logger.info(f"   - File exists before loading: {output_file.exists()}")
            logger.info(f"   - File size before loading: {output_file.stat().st_size if output_file.exists() else 'N/A'}")
            
            wb = openpyxl.load_workbook(str(output_file))
            ws = wb.active
            
            logger.info(f"✅ DATAFILTER: Workbook loaded successfully")
            logger.info(f"   - Worksheet title: {ws.title}")
            logger.info(f"   - Worksheet max_row: {ws.max_row}")
            logger.info(f"   - Worksheet max_column: {ws.max_column}")
            
        except Exception as e:
            logger.error(f"❌ DATAFILTER: Failed to load workbook from {output_file}: {e}")
            logger.error(f"   - File exists: {output_file.exists()}")
            logger.error(f"   - File size: {output_file.stat().st_size if output_file.exists() else 'N/A'}")
            raise TSConverterError(f"Could not open output file for processing: {str(e)}")
        
        # Get initial stats
        initial_rows = ws.max_row
        logger.info(f"Initial rows: {initial_rows}")
        
        try:
            # Step 5.1: Remove NA rows
            na_removed = self.remove_na_rows(ws)
            
            # Step 5.2: Find SD duplicate groups (for later use)
            duplicate_groups = self.find_sd_duplicates(ws)
            
            # Step 5.3: Clear K,L,M for all SD rows
            sd_cleared = self.clear_sd_row_data(ws)
            
            # Step 5.4: Deduplicate SD rows
            sd_removed = self.deduplicate_sd_rows(ws)
            
            # Get final stats
            final_rows = ws.max_row
            total_removed = na_removed + sd_removed
            
            logger.info("Processing Summary:")
            logger.info(f"  Initial rows: {initial_rows}")
            logger.info(f"  NA rows removed: {na_removed}")
            logger.info(f"  SD rows cleared (K,L,M): {sd_cleared}")
            logger.info(f"  SD duplicates removed: {sd_removed}")
            logger.info(f"  Total rows removed: {total_removed}")
            logger.info(f"  Final rows: {final_rows}")
            
            # Save output file with enhanced error handling
            try:
                logger.info(f"💾 DATAFILTER: Saving workbook to: {output_file}")
                logger.info(f"   - Output file before save exists: {output_file.exists()}")
                logger.info(f"   - Output file path: {output_file}")
                logger.info(f"   - Output file absolute path: {output_file.absolute()}")
                logger.info(f"   - Output file parent: {output_file.parent}")
                logger.info(f"   - Output file parent exists: {output_file.parent.exists()}")
                
                wb.save(str(output_file))
                
                logger.info(f"✅ DATAFILTER: Workbook save operation completed")
                logger.info(f"   - Output file after save exists: {output_file.exists()}")
                
                # Verify file was actually saved and is accessible
                if not output_file.exists():
                    logger.error(f"❌ DATAFILTER: Output file was not created successfully!")
                    logger.error(f"   - Expected path: {output_file}")
                    logger.error(f"   - Absolute path: {output_file.absolute()}")
                    logger.error(f"   - Parent directory exists: {output_file.parent.exists()}")
                    
                    # List parent directory contents
                    if output_file.parent.exists():
                        logger.error(f"   - Parent directory contents:")
                        for item in output_file.parent.iterdir():
                            logger.error(f"     - {item}")
                    
                    raise TSConverterError(f"Output file was not created successfully: {output_file}")
                    
                file_size = output_file.stat().st_size
                logger.info(f"✅ DATAFILTER: File verification successful")
                logger.info(f"   - Output file size: {file_size} bytes")
                logger.info(f"   - Output file path: {output_file}")
                
            except Exception as save_error:
                logger.error(f"❌ DATAFILTER: Failed to save file: {save_error}")
                logger.error(f"   - Exception type: {type(save_error).__name__}")
                logger.error(f"   - Exception args: {save_error.args}")
                logger.error(f"   - Output file path: {output_file}")
                logger.error(f"   - Output file exists: {output_file.exists()}")
                logger.error(f"   - Output file parent exists: {output_file.parent.exists()}")
                raise TSConverterError(f"Could not save output file: {str(save_error)}")
                
        except Exception as process_error:
            logger.error(f"❌ DATAFILTER: Processing error during Step 5: {process_error}")
            logger.error(f"   - Exception type: {type(process_error).__name__}")
            logger.error(f"   - Exception args: {process_error.args}")
            raise TSConverterError(f"Data processing failed: {str(process_error)}")
        finally:
            # Always close workbook to free resources
            try:
                logger.info(f"🔒 DATAFILTER: Closing workbook...")
                wb.close()
                logger.info(f"✅ DATAFILTER: Workbook closed successfully")
            except Exception as close_error:
                logger.warning(f"⚠️ DATAFILTER: Could not close workbook properly: {close_error}")
        
        logger.info(f"🎉 DATAFILTER: Process completed successfully!")
        logger.info(f"   - Final output path: {output_file}")
        logger.info(f"   - Final output absolute: {output_file.absolute()}")
        logger.info(f"   - Final output exists: {output_file.exists()}")
        logger.info(f"   - Returning: {str(output_file)}")
        
        return str(output_file)

def main():
    """Command line interface for data filtering and deduplication"""
    parser = argparse.ArgumentParser(description='Data Filter Step 5 - Filter and Deduplicate')
    parser.add_argument('step4_file', help='Step4 input file (*.xlsx)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize filter
    filter_processor = DataFilter(args.base_dir)
    
    try:
        result = filter_processor.process_file(args.step4_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()