#!/usr/bin/env python3
"""
Debug script để investigate merged cells issue trong Step 3
"""

import openpyxl
from openpyxl.utils import get_column_letter

def inspect_source_merged_cells():
    """Inspect merged cells trong source file"""
    print("=== SOURCE FILE MERGED CELLS ANALYSIS ===")
    
    try:
        wb = openpyxl.load_workbook('input/Test plan Purpingla ver 2.xlsx')
        ws = wb['M-Textile ']  # Note the trailing space
        
        print(f"Sheet: {ws.title}")
        print(f"Total merged ranges: {len(ws.merged_cells.ranges)}")
        
        # Check all merged ranges
        j66_j67_ranges = []
        for i, merged_range in enumerate(ws.merged_cells.ranges):
            print(f"\nMerged Range #{i+1}: {merged_range}")
            print(f"  Rows: {merged_range.min_row} to {merged_range.max_row}")
            print(f"  Cols: {merged_range.min_col} to {merged_range.max_col} ({get_column_letter(merged_range.min_col)} to {get_column_letter(merged_range.max_col)})")
            
            # Check if affects rows 66-67
            if merged_range.min_row <= 67 and merged_range.max_row >= 66:
                print(f"  ⚠️  AFFECTS ROWS 66-67!")
                j66_j67_ranges.append(merged_range)
                
            # Check if affects column J (column 10)
            if merged_range.min_col <= 10 <= merged_range.max_col:
                print(f"  ⚠️  AFFECTS COLUMN J!")
                
            # Get content from top-left cell
            top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            print(f"  Content: '{top_left_cell.value}'")
        
        print(f"\n=== RANGES AFFECTING ROWS 66-67: {len(j66_j67_ranges)} ===")
        for rng in j66_j67_ranges:
            print(f"  {rng}")
            
        wb.close()
        
    except Exception as e:
        print(f"Error inspecting source file: {e}")

def inspect_step3_merged_cells():
    """Inspect merged cells trong Step3 output"""
    print("\n=== STEP3 OUTPUT MERGED CELLS ANALYSIS ===")
    
    try:
        wb = openpyxl.load_workbook('output/Test plan Purpingla ver 2 - Step3.xlsx')
        ws = wb['M-Textile ']
        
        print(f"Sheet: {ws.title}")
        print(f"Total merged ranges: {len(ws.merged_cells.ranges)}")
        
        j66_j67_ranges = []
        for i, merged_range in enumerate(ws.merged_cells.ranges):
            print(f"\nMerged Range #{i+1}: {merged_range}")
            print(f"  Rows: {merged_range.min_row} to {merged_range.max_row}")
            print(f"  Cols: {merged_range.min_col} to {merged_range.max_col} ({get_column_letter(merged_range.min_col)} to {get_column_letter(merged_range.max_col)})")
            
            # Special focus on J66, J67 area
            if (merged_range.min_row <= 67 and merged_range.max_row >= 66 and
                merged_range.min_col <= 10 <= merged_range.max_col):
                print(f"  🎯 FOUND: This range includes J66-J67!")
                j66_j67_ranges.append(merged_range)
                
            # Get content from top-left cell
            top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            print(f"  Content: '{top_left_cell.value}'")
        
        print(f"\n=== RANGES AFFECTING J66-J67: {len(j66_j67_ranges)} ===")
        for rng in j66_j67_ranges:
            print(f"  {rng}")
            
        wb.close()
        
    except Exception as e:
        print(f"Error inspecting Step3 file: {e}")

def compare_specific_area():
    """Cell-by-cell comparison khu vực J66, J67"""
    print("\n=== CELL-BY-CELL COMPARISON (Rows 65-68, Cols I-L) ===")
    
    try:
        source_wb = openpyxl.load_workbook('input/Test plan Purpingla ver 2.xlsx')
        step3_wb = openpyxl.load_workbook('output/Test plan Purpingla ver 2 - Step3.xlsx')
        
        source_ws = source_wb['M-Textile ']
        step3_ws = step3_wb['M-Textile ']
        
        for row in range(65, 69):  # 65-68
            print(f"\n--- ROW {row} ---")
            for col in range(9, 13):  # I-L (9-12)
                col_letter = get_column_letter(col)
                
                source_cell = source_ws.cell(row, col)
                step3_cell = step3_ws.cell(row, col)
                
                print(f"\n{col_letter}{row}:")
                print(f"  Source: '{source_cell.value}' (Type: {type(source_cell).__name__})")
                print(f"  Step3:  '{step3_cell.value}' (Type: {type(step3_cell).__name__})")
                
                # Check if cell is part of merged range
                source_merged = any(rng for rng in source_ws.merged_cells.ranges 
                                  if row >= rng.min_row and row <= rng.max_row and 
                                     col >= rng.min_col and col <= rng.max_col)
                step3_merged = any(rng for rng in step3_ws.merged_cells.ranges 
                                 if row >= rng.min_row and row <= rng.max_row and 
                                    col >= rng.min_col and col <= rng.max_col)
                
                print(f"  Source merged: {source_merged}")
                print(f"  Step3 merged:  {step3_merged}")
                
                # Show difference if any
                if source_cell.value != step3_cell.value:
                    print(f"  🔴 VALUE CHANGED!")
                if source_merged != step3_merged:
                    print(f"  🔴 MERGE STATUS CHANGED!")
        
        source_wb.close()
        step3_wb.close()
        
    except Exception as e:
        print(f"Error comparing files: {e}")

def analyze_fill_operations():
    """Analyze fill operations từ logs"""
    print("\n=== ANALYZING STEP 3 FILL OPERATIONS ===")
    print("From previous logs, we know:")
    print("- M-Textile filled 1,044 cells (J:348, K:348, L:348)")
    print("- Product combination found at row 17")
    print("- Data processing starts at row 19 (17 + 2)")
    print("- Rows 66-67 would be processed (within range)")
    
    # Calculate expected position
    print("\nCalculated positions:")
    print("- Product combination at row 17")
    print("- Data starts at row 19")  
    print("- Rows 66-67 are data rows 48-49 of processing")
    print("- This is well within the fill range")
    print("- Expected: J66, J67 should be filled with material data")

def verify_theories():
    """Verify các theories about merged cell origin"""
    print("\n=== THEORY VERIFICATION ===")
    
    theories = [
        "Theory 1: Source has cross-column merged cells (I66:K67 or similar)",
        "Theory 2: Step 3 fill operation created merged cells somehow", 
        "Theory 3: Excel formatting preservation during copy/load/save",
        "Theory 4: Openpyxl copy operation side effect"
    ]
    
    print("Based on inspection results above:")
    print("\nTheories to verify:")
    for i, theory in enumerate(theories, 1):
        print(f"{i}. {theory}")

def main():
    """Main debug function"""
    print("🔍 DEBUG: Merged Cells Investigation - Test plan Purpingla ver 2")
    print("=" * 80)
    
    # Step 1: Inspect source file
    inspect_source_merged_cells()
    
    # Step 2: Inspect Step3 output
    inspect_step3_merged_cells()
    
    # Step 3: Compare specific area
    compare_specific_area()
    
    # Step 4: Analyze fill operations
    analyze_fill_operations()
    
    # Step 5: Theory verification summary
    verify_theories()
    
    print("\n" + "=" * 80)
    print("🔍 DEBUG INVESTIGATION COMPLETE")

if __name__ == "__main__":
    main()