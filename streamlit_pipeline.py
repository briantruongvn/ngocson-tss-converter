"""
Streamlit Pipeline Integration for TSS Converter
Wraps the existing 6-step pipeline with progress tracking and error handling.
"""

import os
import sys
import tempfile
import shutil
import time
from pathlib import Path
from typing import Dict, Any, Optional, Callable, Tuple
import logging
import traceback

# Import existing pipeline modules (they use function-based approach)
import step1_template_creation
import step2_data_extraction
import step3_pre_mapping_fill
import step4_data_mapping
import step5_filter_deduplicate
import step6_article_crossref
from common.exceptions import TSConverterError
from common.validation import FileValidator
from common.quality_reporter import get_global_reporter, reset_global_reporter
from common.error_handler import global_error_handler
from common.security import FileValidator as SecurityFileValidator, validate_path_security, sanitize_filename, generate_secure_filename, SecurityError
from common.session_manager import session_manager, ProcessingState, safe_update_session_state, safe_get_session_value
from config_streamlit import get_temp_directory, STREAMLIT_CONFIG

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ResourceManager:
    """Context manager for handling temporary files and cleanup with security validation"""
    
    def __init__(self, allowed_base_dir: Optional[Path] = None):
        self.temp_files = []
        self.allowed_base_dir = allowed_base_dir or get_temp_directory()
        
    def __enter__(self):
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        # Cleanup temporary files with path validation
        for file_path in self.temp_files:
            try:
                if isinstance(file_path, (str, Path)):
                    path_obj = Path(file_path)
                    
                    # Security check: validate path is within allowed base
                    if not validate_path_security(path_obj, self.allowed_base_dir):
                        logger.warning(f"Skipping cleanup of suspicious path: {path_obj}")
                        continue
                    
                    if path_obj.exists():
                        if path_obj.is_file():
                            path_obj.unlink(missing_ok=True)
                        elif path_obj.is_dir():
                            shutil.rmtree(path_obj, ignore_errors=True)
                            
            except (OSError, PermissionError) as e:
                logger.warning(f"Permission error cleaning up {file_path}: {e}")
            except Exception as e:
                logger.warning(f"Failed to cleanup temp file {file_path}: {e}")
                
    def add_temp_file(self, file_path):
        """Add a file or directory to be cleaned up with security validation"""
        try:
            path_obj = Path(file_path)
            
            # Security check: validate path is within allowed base
            if validate_path_security(path_obj, self.allowed_base_dir):
                self.temp_files.append(file_path)
            else:
                logger.warning(f"Refusing to track suspicious path: {file_path}")
                
        except Exception as e:
            logger.warning(f"Failed to add temp file {file_path}: {e}")

def with_retry(max_retries=3, exceptions=(Exception,), backoff_factor=0.5):
    """Decorator for retrying functions on failure"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except exceptions as e:
                    if attempt == max_retries:
                        raise e
                    wait_time = backoff_factor * (2 ** attempt)
                    time.sleep(wait_time)
                    logger.warning(f"Retrying {func.__name__} (attempt {attempt + 1}/{max_retries + 1}) after {wait_time}s")
            return func(*args, **kwargs)
        return wrapper
    return decorator

class ProgressCallback:
    """Callback class for tracking pipeline progress"""
    
    def __init__(self, update_func: Optional[Callable] = None):
        self.update_func = update_func
        self.current_step = 0
        self.step_status = {f"step{i}": "pending" for i in range(1, 7)}
        self.start_time = time.time()
        
    def start_step(self, step_num: int, step_name: str):
        """Mark step as started"""
        self.current_step = step_num
        self.step_status[f"step{step_num}"] = "running"
        
        if self.update_func:
            self.update_func({
                "current_step": step_num,
                "step_status": self.step_status.copy(),
                "message": f"Running {step_name}..."
            })
    
    def complete_step(self, step_num: int, step_name: str):
        """Mark step as completed"""
        self.step_status[f"step{step_num}"] = "completed"
        
        if self.update_func:
            self.update_func({
                "current_step": step_num,
                "step_status": self.step_status.copy(),
                "message": f"Completed {step_name}"
            })
    
    def error_step(self, step_num: int, error_message: str):
        """Mark step as error"""
        self.step_status[f"step{step_num}"] = "error"
        
        if self.update_func:
            self.update_func({
                "current_step": step_num,
                "step_status": self.step_status.copy(),
                "message": f"Error: {error_message}",
                "error": True
            })

class StreamlitTSSPipeline:
    """
    Streamlit wrapper for TSS Converter pipeline
    Provides progress tracking, file management, and error handling for web interface with security features
    """
    
    def __init__(self, temp_dir: Optional[Path] = None):
        self.temp_dir = temp_dir or get_temp_directory()
        self.current_session_id = None
        self.processing_stats = {}
        
        # Initialize security validator with configuration from Streamlit settings
        from config_streamlit import get_validation_config
        validation_config = get_validation_config()
        self.security_validator = SecurityFileValidator(
            max_size=validation_config.get('max_file_size_mb', 50) * 1024 * 1024,
            strict_mode=validation_config.get('strict_mode', False),
            enable_fallbacks=validation_config.get('enable_fallbacks', True)
        )
        
        # Ensure temp directory is secure
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize session manager
        session_manager.initialize_session_state()
    
    def _validate_paths_security(self, *paths: Path) -> None:
        """Helper method to validate multiple paths for security"""
        for path in paths:
            if not validate_path_security(path, self.temp_dir):
                raise SecurityError(f"Path validation failed for {path}")
    
    def _handle_cli_output_file(self, cli_output_path: str, output_dir: Path, 
                               step_name: str = "", secure_permissions: bool = True) -> Path:
        """
        Helper method to handle CLI output file with security validation and proper session management
        
        Args:
            cli_output_path: Path returned from CLI module
            output_dir: Session output directory
            step_name: Step name for logging (optional)
            secure_permissions: Whether to set secure file permissions
            
        Returns:
            Path to file in session output directory
        """
        try:
            # Convert CLI output to Path object
            cli_path = Path(cli_output_path)
            
            # Security validation for CLI generated path
            if not validate_path_security(cli_path, Path.cwd()):
                raise SecurityError(f"CLI generated output path validation failed: {cli_path}")
            
            # Generate session output path
            session_output = output_dir / cli_path.name
            
            # Security validation for session output path  
            if not validate_path_security(session_output, self.temp_dir):
                raise SecurityError(f"Session output path validation failed: {session_output}")
            
            # Move CLI output to session directory
            if cli_path != session_output:
                shutil.move(str(cli_path), str(session_output))
                logger.info(f"Moved {step_name} output to session location: {cli_path} -> {session_output}")
            
            # Set secure file permissions if requested
            if secure_permissions:
                session_output.chmod(0o600)
            
            logger.info(f"{step_name} completed successfully: {session_output}")
            return session_output
            
        except Exception as e:
            raise TSConverterError(f"Output file handling failed for {step_name}: {str(e)}")
    
    def _call_cli_with_explicit_output(self, cli_instance, input_file: Path, output_dir: Path,
                                     output_filename: str, step_name: str) -> Path:
        """
        Helper method for CLI calls that need explicit output path (Step 5, 6)
        
        Args:
            cli_instance: CLI module instance to call
            input_file: Input file path
            output_dir: Session output directory
            output_filename: Explicit output filename
            step_name: Step name for logging
            
        Returns:
            Path to file in session output directory
        """
        try:
            # Security validation
            self._validate_paths_security(input_file, output_dir)
            
            # Create explicit session output path
            session_output = output_dir / output_filename
            if not validate_path_security(session_output, self.temp_dir):
                raise SecurityError(f"Session output path validation failed: {session_output}")
            
            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Call CLI with explicit output path
            cli_result = cli_instance.process_file(str(input_file), str(session_output))
            
            # Verify result (CLI should have created file at session_output)
            result_path = Path(cli_result)
            if not result_path.exists():
                raise TSConverterError(f"CLI claimed success but file not found: {result_path}")
            
            # Set secure permissions
            result_path.chmod(0o600)
            
            logger.info(f"{step_name} completed successfully: {result_path}")
            return result_path
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"{step_name} failed: {str(e)}")
    
    def _call_data_mapper_cli(self, source_file: Path, step3_output: Path, output_dir: Path,
                             output_filename: str) -> Path:
        """
        Helper method for Step 4 DataMapper - handles complex file dependencies
        
        Step 4 DataMapper needs access to both source file and Step3 output:
        - source_file: Used for reference and compatibility (Step 4 actually uses Step3 data)
        - step3_output: The actual input file containing data to be mapped
        
        Args:
            source_file: Original input file (kept for compatibility, Step 4 uses step3_output)
            step3_output: Step3 output file containing source data to be mapped
            output_dir: Session output directory
            output_filename: Target output filename
            
        Returns:
            Path to Step4 file in session directory
        """
        try:
            # Security validation
            self._validate_paths_security(source_file, step3_output, output_dir)
            
            # Create session output path
            session_output = output_dir / output_filename
            if not validate_path_security(session_output, self.temp_dir):
                raise SecurityError(f"Session output path validation failed: {session_output}")
            
            # DataMapper auto-detects Step2 template, so we need to ensure Step3 is accessible
            # Copy step3_output to expected location if needed
            step3_name = step3_output.name
            expected_step3_path = output_dir / step3_name
            temp_step3_created = False
            
            if step3_output != expected_step3_path and not expected_step3_path.exists():
                shutil.copy2(str(step3_output), str(expected_step3_path))
                temp_step3_created = True
                logger.info(f"Copied Step3 file for processing: {step3_output} -> {expected_step3_path}")
            
            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Direct CLI module call - Single source of truth!
            # Step 4 DataMapper expects Step3 output as input (not source_file)
            parent_dir = output_dir.parent  # Avoid double output directory
            mapper = step4_data_mapping.DataMapper(base_dir=str(parent_dir))
            cli_result = mapper.process_file(str(step3_output), str(session_output))
            
            # Clean up temporary file if created
            if temp_step3_created and expected_step3_path.exists():
                expected_step3_path.unlink(missing_ok=True)
                logger.info(f"Cleaned up temporary Step3 file: {expected_step3_path}")
            
            # Verify and return result
            result_path = Path(cli_result)
            if not result_path.exists():
                raise TSConverterError(f"DataMapper claimed success but file not found: {result_path}")
            
            # Ensure output is at session location
            if result_path != session_output:
                shutil.move(str(result_path), str(session_output))
                logger.info(f"Moved Step4 output to session location: {result_path} -> {session_output}")
            
            # Set secure permissions
            session_output.chmod(0o600)
            
            logger.info(f"Step 4 (Data Mapping) completed successfully: {session_output}")
            return session_output
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 4 (Data Mapping) failed: {str(e)}")
        
    def create_session_directory(self) -> Path:
        """Create unique session directory for file processing with security validation"""
        try:
            # Generate secure session ID
            import random
            import string
            random_part = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
            session_id = f"session_{int(time.time())}_{random_part}"
            self.current_session_id = session_id
            
            session_dir = self.temp_dir / session_id
            
            # Security check: validate session directory is within temp_dir
            if not validate_path_security(session_dir, self.temp_dir):
                raise SecurityError("Session directory path validation failed")
            
            session_dir.mkdir(parents=True, exist_ok=True)
            
            # Create subdirectories with restricted permissions
            input_dir = session_dir / "input"
            output_dir = session_dir / "output"
            
            input_dir.mkdir(mode=0o700, exist_ok=True)
            output_dir.mkdir(mode=0o700, exist_ok=True)
            
            # Update session state safely
            safe_update_session_state({
                'session_id': session_id,
                'session_dir': str(session_dir)
            })
            
            logger.info(f"Created secure session directory: {session_dir}")
            return session_dir
            
        except Exception as e:
            logger.error(f"Failed to create session directory: {e}")
            raise SecurityError(f"Session creation failed: {str(e)}")
    
    def save_uploaded_file(self, file_data: bytes, filename: str) -> Path:
        """Save uploaded file to session directory with comprehensive security validation"""
        try:
            # Step 1: Validate file data using security validator
            is_valid, error_msg = self.security_validator.validate_file(file_data, filename)
            if not is_valid:
                raise SecurityError(f"File security validation failed: {error_msg}")
            
            # Step 2: Sanitize filename
            safe_filename = sanitize_filename(filename)
            if safe_filename != filename:
                logger.info(f"Filename sanitized: {filename} -> {safe_filename}")
            
            # Step 3: Create session directory if needed
            if not self.current_session_id:
                self.create_session_directory()
            
            session_dir = self.temp_dir / self.current_session_id
            input_file_path = session_dir / "input" / safe_filename
            
            # Step 4: Security check - validate final path
            if not validate_path_security(input_file_path, self.temp_dir):
                raise SecurityError("File path validation failed - potential path traversal")
            
            # Step 5: Check file size before writing
            max_size = STREAMLIT_CONFIG.get("max_file_size_mb", 50) * 1024 * 1024
            if len(file_data) > max_size:
                raise SecurityError(f"File too large: {len(file_data)} bytes > {max_size} bytes")
            
            # Step 6: Write file securely with restricted permissions
            input_file_path.parent.mkdir(parents=True, exist_ok=True)
            with open(input_file_path, "wb") as f:
                f.write(file_data)
            
            # Set restrictive file permissions
            input_file_path.chmod(0o600)
            
            # Step 7: Update session state safely
            safe_update_session_state({
                'uploaded_file_info': {
                    'original_filename': filename,
                    'safe_filename': safe_filename,
                    'file_path': str(input_file_path),
                    'file_size': len(file_data),
                    'upload_time': time.time()
                }
            })
            
            logger.info(f"Securely saved uploaded file to: {input_file_path}")
            return input_file_path
            
        except SecurityError:
            raise
        except Exception as e:
            logger.error(f"Failed to save uploaded file: {e}")
            raise SecurityError(f"File save operation failed: {str(e)}")
    
    def process_pipeline(self, 
                        input_file_path: Path, 
                        progress_callback: Optional[ProgressCallback] = None) -> Tuple[bool, Path, Dict[str, Any]]:
        """
        Run complete 6-step pipeline with progress tracking and security validation
        
        Args:
            input_file_path: Path to input Excel file
            progress_callback: Callback for progress updates
            
        Returns:
            Tuple of (success, output_file_path, processing_stats)
        """
        start_time = time.time()
        
        try:
            # Security validation: verify input path is safe
            if not validate_path_security(input_file_path, self.temp_dir):
                raise SecurityError("Input file path validation failed")
            
            session_dir = input_file_path.parent.parent
            output_dir = session_dir / "output"
            
            # Security validation: verify output directory is safe
            if not validate_path_security(output_dir, self.temp_dir):
                raise SecurityError("Output directory path validation failed")
            
            # Update processing state safely
            session_manager.update_processing_state(ProcessingState.PROCESSING)
            
            # Reset and initialize quality reporter
            reset_global_reporter()
            reporter = get_global_reporter()
            reporter.start_processing()
            
            # Initialize processing stats
            self.processing_stats = {
                "start_time": start_time,
                "input_file": str(input_file_path.name),
                "steps_completed": 0,
                "errors": []
            }
            
            # Update session state with processing info
            safe_update_session_state({
                'processing_start_time': start_time,
                'processing_stats': self.processing_stats
            })
            
            # Step 1: Template Creation
            if progress_callback:
                progress_callback.start_step(1, "Create Template")
            
            step1_output = self._run_step1(input_file_path, output_dir)
            
            if progress_callback:
                progress_callback.complete_step(1, "Create Template")
            self.processing_stats["steps_completed"] = 1
            
            # Step 2: Data Extraction
            if progress_callback:
                progress_callback.start_step(2, "Extract Data")
            
            step2_output = self._run_step2(step1_output, input_file_path, output_dir)
            
            if progress_callback:
                progress_callback.complete_step(2, "Extract Data")
            self.processing_stats["steps_completed"] = 2
            
            # Step 3: Pre-mapping Fill
            if progress_callback:
                progress_callback.start_step(3, "Pre-mapping Fill")
            
            step3_output = self._run_step3(step2_output, output_dir)
            
            if progress_callback:
                progress_callback.complete_step(3, "Pre-mapping Fill")
            self.processing_stats["steps_completed"] = 3
            
            # Step 4: Data Mapping
            if progress_callback:
                progress_callback.start_step(4, "Data Mapping")
            
            step4_output = self._run_step4(input_file_path, step3_output, output_dir)
            
            if progress_callback:
                progress_callback.complete_step(4, "Data Mapping")
            self.processing_stats["steps_completed"] = 4
            
            # Step 5: Filter & Deduplicate
            if progress_callback:
                progress_callback.start_step(5, "Filter & Deduplicate")
            
            logger.info(f"Starting Step 5 with input: {step4_output}")
            logger.info(f"Step 5 output directory: {output_dir}")
            
            step5_output = self._run_step5(step4_output, output_dir)
            
            logger.info(f"Step 5 output: {step5_output}")
            
            if progress_callback:
                progress_callback.complete_step(5, "Filter & Deduplicate")
            self.processing_stats["steps_completed"] = 5
            
            # Step 6: Article Cross-Reference
            if progress_callback:
                progress_callback.start_step(6, "Article Cross-Reference")
            
            logger.info(f"Starting Step 6 with input: {step5_output}")
            logger.info(f"Step 6 output directory: {output_dir}")
            
            final_output = self._run_step6(step5_output, output_dir)
            
            logger.info(f"Step 6 final output: {final_output}")
            
            if progress_callback:
                progress_callback.complete_step(6, "Article Cross-Reference")
            self.processing_stats["steps_completed"] = 6
            
            # Calculate final statistics
            end_time = time.time()
            reporter.end_processing()
            
            # Get quality summary
            quality_summary = reporter.get_user_summary()
            
            # Update processing state
            session_manager.update_processing_state(ProcessingState.COMPLETED)
            
            self.processing_stats.update({
                "end_time": end_time,
                "processing_time": end_time - start_time,
                "success": True,
                "final_output": str(final_output),
                "quality_score": quality_summary["quality_score"],
                "warnings_count": quality_summary["warnings_count"],
                "errors_count": quality_summary["errors_count"],
                "quality_summary": quality_summary
            })
            
            # Update session state with final results
            logger.info(f"🔄 PIPELINE: Updating final session state...")
            logger.info(f"   - final_output: {final_output}")
            logger.info(f"   - final_output exists before session update: {final_output.exists()}")
            
            safe_update_session_state({
                'processing_stats': self.processing_stats,
                'output_file_path': str(final_output),
                'processing_complete': True
            })
            
            logger.info(f"✅ PIPELINE: Session state updated successfully")
            logger.info(f"   - final_output exists after session update: {final_output.exists()}")
            
            logger.info(f"🎉 Pipeline completed successfully: {final_output}")
            logger.info(f"📊 Quality score: {quality_summary['quality_score']:.1f}/100")
            
            logger.info(f"📈 PIPELINE: Preparing success stats...")
            logger.info(f"   - final_output exists before stats prep: {final_output.exists()}")
            
            # Add success metrics to error handler
            success_stats = self.processing_stats.copy()
            success_stats['error_summary'] = global_error_handler.get_error_summary()
            
            logger.info(f"✅ PIPELINE: Success stats prepared")
            logger.info(f"   - final_output exists before return: {final_output.exists()}")
            
            # CRITICAL DEBUG: Add extensive pre-return validation
            logger.info(f"🔍 FINAL VALIDATION before return:")
            logger.info(f"   - final_output type: {type(final_output)}")
            logger.info(f"   - final_output: {final_output}")
            logger.info(f"   - final_output absolute: {final_output.absolute()}")
            logger.info(f"   - final_output exists: {final_output.exists()}")
            logger.info(f"   - final_output parent: {final_output.parent}")
            logger.info(f"   - final_output parent exists: {final_output.parent.exists()}")
            
            if final_output.exists():
                try:
                    file_size = final_output.stat().st_size
                    logger.info(f"   - final_output size: {file_size} bytes")
                    
                    # Try to read file to ensure it's accessible
                    with open(final_output, 'rb') as f:
                        first_bytes = f.read(100)
                    logger.info(f"   - final_output readable: True (first 100 bytes: {len(first_bytes)})")
                    
                except Exception as file_check_error:
                    logger.error(f"❌ FINAL VALIDATION: File check failed: {file_check_error}")
            else:
                logger.error(f"❌ FINAL VALIDATION: File does not exist before return!")
            
            logger.info(f"🚀 PIPELINE: Returning success result...")
            
            # Add extra file existence check right before return
            if not final_output.exists():
                logger.error(f"🚨 CRITICAL: File disappeared right before return!")
                logger.error(f"   - Expected path: {final_output}")
                logger.error(f"   - Directory contents:")
                if final_output.parent.exists():
                    for item in final_output.parent.iterdir():
                        logger.error(f"     - {item}")
            
            return True, final_output, success_stats
            
        except SecurityError as se:
            error_msg = f"Security error: {str(se)}"
            logger.error(f"Security violation during pipeline: {error_msg}")
            
            # Update processing state safely
            session_manager.update_processing_state(ProcessingState.ERROR)
            
            if progress_callback:
                current_step = self.processing_stats.get("steps_completed", 0) + 1
                progress_callback.error_step(current_step, error_msg)
            
            self.processing_stats.update({
                "end_time": time.time(),
                "processing_time": time.time() - start_time,
                "success": False,
                "error_message": error_msg,
                "error_type": "security_error"
            })
            
            safe_update_session_state({
                'processing_stats': self.processing_stats,
                'error_message': error_msg
            })
            
            return False, None, self.processing_stats
            
        except Exception as e:
            error_msg = str(e)
            error_details = traceback.format_exc()
            
            logger.error(f"Pipeline failed: {error_msg}")
            logger.error(f"Error details: {error_details}")
            
            # Update processing state safely
            session_manager.update_processing_state(ProcessingState.ERROR)
            
            if progress_callback:
                current_step = self.processing_stats.get("steps_completed", 0) + 1
                progress_callback.error_step(current_step, error_msg)
            
            self.processing_stats.update({
                "end_time": time.time(),
                "processing_time": time.time() - start_time,
                "success": False,
                "error_message": error_msg,
                "error_details": error_details if STREAMLIT_CONFIG.get("show_error_details") else None
            })
            
            safe_update_session_state({
                'processing_stats': self.processing_stats,
                'error_message': error_msg
            })
            
            return False, None, self.processing_stats
    
    def _run_step1(self, input_file: Path, output_dir: Path) -> Path:
        """Run Step 1: Template Creation - Direct CLI module call with security wrapper"""
        try:
            # Security validation using helper
            self._validate_paths_security(input_file, output_dir)
            
            # Direct CLI module call - Single source of truth!
            creator = step1_template_creation.TemplateCreator()
            cli_output = creator.create_template(str(input_file))
            
            # Handle output file using helper
            return self._handle_cli_output_file(cli_output, output_dir, "Step 1 (Template Creation)")
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 1 failed: {str(e)}")
    
    def _run_step2(self, step1_output: Path, source_file: Path, output_dir: Path) -> Path:
        """Run Step 2: Data Extraction - Direct CLI module call with security wrapper"""
        try:
            # Security validation using helper
            self._validate_paths_security(step1_output, source_file, output_dir)
            
            # Create Step2 output filename
            output_filename = step1_output.name.replace(" - Step1.xlsx", " - Step2.xlsx")
            if not output_filename.endswith(" - Step2.xlsx"):
                output_filename = step1_output.stem + " - Step2.xlsx"
            
            # Create explicit session output path
            session_output = output_dir / output_filename
            if not validate_path_security(session_output, self.temp_dir):
                raise SecurityError(f"Session output path validation failed: {session_output}")
            
            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Direct CLI module call with explicit output - Single source of truth!
            extractor = step2_data_extraction.DataExtractor()
            logger.info(f"Using M-Textile extraction logic for Step 2")
            cli_result = extractor.process_m_textile_file(
                str(step1_output), 
                str(source_file),
                str(session_output)
            )
            
            # Verify result
            result_path = Path(cli_result)
            if not result_path.exists():
                raise TSConverterError(f"DataExtractor claimed success but file not found: {result_path}")
            
            # Set secure permissions
            result_path.chmod(0o600)
            
            logger.info(f"Step 2 (Data Extraction) completed successfully: {result_path}")
            return result_path
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 2 failed: {str(e)}")
    
    def _run_step3(self, step2_output: Path, output_dir: Path) -> Path:
        """Run Step 3: Pre-mapping Fill - Direct CLI module call with security wrapper"""
        try:
            # Security validation using helper
            self._validate_paths_security(step2_output, output_dir)
            
            # Direct CLI module call - Single source of truth!
            filler = step3_pre_mapping_fill.PreMappingFiller()
            cli_output = filler.process_file(str(step2_output))
            
            # Handle output file using helper
            return self._handle_cli_output_file(cli_output, output_dir, "Step 3 (Pre-mapping Fill)")
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 3 failed: {str(e)}")
    
    def _run_step4(self, source_file: Path, step3_output: Path, output_dir: Path) -> Path:
        """
        Run Step 4: Data Mapping - Direct CLI module call with security wrapper
        
        Step 4 maps data from Step3 output (source file with filled data) to Step2 template.
        Note: source_file parameter kept for compatibility but Step 4 actually uses step3_output.
        
        Args:
            source_file: Original input file (kept for interface compatibility)
            step3_output: Step3 output file containing source data with filled information
            output_dir: Session output directory
            
        Returns:
            Path to Step4 output file
        """
        try:
            # Create Step4 output filename
            output_filename = step3_output.name.replace(" - Step3.xlsx", " - Step4.xlsx")
            if not output_filename.endswith(" - Step4.xlsx"):
                # Fallback if naming doesn't match expected pattern
                output_filename = step3_output.stem + " - Step4.xlsx"
            
            # Direct CLI module call using specialized helper - Single source of truth!
            return self._call_data_mapper_cli(source_file, step3_output, output_dir, output_filename)
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 4 failed: {str(e)}")
    
    def _run_step5(self, step4_output: Path, output_dir: Path) -> Path:
        """Run Step 5: Filter & Deduplicate - Direct CLI module call with security wrapper"""
        try:
            # Create Step5 output filename
            output_filename = step4_output.name.replace(" - Step4.xlsx", " - Step5.xlsx")
            if not output_filename.endswith(" - Step5.xlsx"):
                # Fallback if naming doesn't match expected pattern
                output_filename = step4_output.stem + " - Step5.xlsx"
            
            # Direct CLI module call using helper - Single source of truth!
            filter_dedup = step5_filter_deduplicate.DataFilter()
            return self._call_cli_with_explicit_output(
                filter_dedup, step4_output, output_dir, output_filename,
                "Step 5 (Filter & Deduplicate)"
            )
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 5 failed: {str(e)}")
    
    def _run_step6(self, step5_output: Path, output_dir: Path) -> Path:
        """Run Step 6: Article Cross-Reference - Direct CLI module call with security wrapper"""
        try:
            # Create Step6 output with descriptive name
            base_name = step5_output.stem.replace(" - Step5", "")
            output_filename = f"Standard Internal TSS - {base_name}.xlsx"
            
            # Direct CLI module call using helper - Single source of truth!
            crossref = step6_article_crossref.ArticleCrossReference()
            return self._call_cli_with_explicit_output(
                crossref, step5_output, output_dir, output_filename,
                "Step 6 (Article Cross-Reference)"
            )
            
        except SecurityError:
            raise
        except Exception as e:
            raise TSConverterError(f"Step 6 failed: {str(e)}")
    
    def _extract_step5_stats(self, output_file: Path):
        """Extract statistics from Step 5 output for display with security validation"""
        try:
            logger.info(f"📊 EXTRACT_STATS: Starting stats extraction for: {output_file}")
            logger.info(f"   - File exists before extraction: {output_file.exists()}")
            logger.info(f"   - File absolute path: {output_file.absolute()}")
            
            # Security validation: verify file path
            if not validate_path_security(output_file, self.temp_dir):
                logger.warning(f"⚠️ EXTRACT_STATS: Skipping stats extraction for suspicious file: {output_file}")
                return
            
            logger.info(f"✅ EXTRACT_STATS: Security validation passed")
                
            import openpyxl
            logger.info(f"📊 EXTRACT_STATS: Loading workbook from: {output_file}")
            logger.info(f"   - File exists before openpyxl.load: {output_file.exists()}")
            
            wb = openpyxl.load_workbook(output_file, read_only=True)
            ws = wb.active
            
            logger.info(f"📊 EXTRACT_STATS: Workbook loaded successfully")
            logger.info(f"   - Worksheet title: {ws.title}")
            logger.info(f"   - Max rows: {ws.max_row}")
            
            # Count final rows (excluding header)
            final_rows = max(0, ws.max_row - 3)  # Subtract header rows, ensure non-negative
            
            # Update stats safely
            stats_update = {
                "final_rows": final_rows,
            }
            
            logger.info(f"📊 EXTRACT_STATS: Updating processing stats")
            logger.info(f"   - Final rows: {final_rows}")
            
            self.processing_stats.update(stats_update)
            
            # Also update session state
            logger.info(f"📊 EXTRACT_STATS: Updating session state")
            safe_update_session_state({
                'processing_stats': self.processing_stats
            })
            
            logger.info(f"📊 EXTRACT_STATS: Closing workbook")
            wb.close()
            logger.info(f"✅ EXTRACT_STATS: Stats extraction completed successfully")
            
        except Exception as e:
            logger.error(f"❌ EXTRACT_STATS: Stats extraction failed: {e}")
            logger.error(f"   - Exception type: {type(e).__name__}")
            logger.error(f"   - Exception args: {e.args}")
            logger.error(f"   - File path: {output_file}")
            logger.error(f"   - File exists: {output_file.exists() if hasattr(output_file, 'exists') else 'N/A'}")
            
            # Import traceback for full error context
            import traceback
            logger.error(f"   - Full traceback:\n{traceback.format_exc()}")
            
            # Don't re-raise, just log the warning
            logger.warning(f"⚠️ EXTRACT_STATS: Could not extract statistics: {e}")
    
    def cleanup_session(self):
        """Clean up session directory and temporary files with security validation"""
        if self.current_session_id:
            session_dir = self.temp_dir / self.current_session_id
            try:
                # Security check: validate session directory path
                if not validate_path_security(session_dir, self.temp_dir):
                    logger.warning(f"Skipping cleanup of suspicious session path: {session_dir}")
                    return
                
                if session_dir.exists():
                    # Secure cleanup: iterate through files and validate each path
                    for item in session_dir.rglob('*'):
                        try:
                            if validate_path_security(item, self.temp_dir):
                                if item.is_file():
                                    item.unlink(missing_ok=True)
                                elif item.is_dir():
                                    # Only remove if empty
                                    try:
                                        item.rmdir()
                                    except OSError:
                                        pass  # Directory not empty
                        except Exception as cleanup_error:
                            logger.warning(f"Failed to cleanup item {item}: {cleanup_error}")
                    
                    # Finally remove session directory if empty
                    try:
                        session_dir.rmdir()
                        logger.info(f"Cleaned up session: {self.current_session_id}")
                    except OSError:
                        logger.warning(f"Session directory {session_dir} not empty after cleanup")
                        
            except (OSError, PermissionError) as e:
                logger.error(f"Permission error during session cleanup {self.current_session_id}: {e}")
            except Exception as e:
                logger.error(f"Failed to cleanup session {self.current_session_id}: {e}")
            finally:
                self.current_session_id = None
                session_manager.cleanup_session_state()
    
    def validate_input_file(self, file_path: Path) -> Tuple[bool, str]:
        """
        Validate input file format and structure with enhanced error handling and graceful degradation
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        validation_warnings = []
        
        try:
            # Security validation: check path security
            if not validate_path_security(file_path, self.temp_dir):
                logger.error(f"Path security validation failed for: {file_path}")
                return False, "File path security validation failed"
            
            # Basic file validation
            if not file_path.exists():
                return False, "File không tồn tại"
            
            if not file_path.suffix.lower() == '.xlsx':
                return False, "File phải có định dạng .xlsx"
            
            # File size validation with grace period
            try:
                file_size = file_path.stat().st_size
                file_size_mb = file_size / (1024 * 1024)
                max_size = STREAMLIT_CONFIG.get("max_file_size_mb", 50)
                
                if file_size_mb > max_size * 2:  # Hard limit
                    return False, f"File quá lớn (>{max_size*2}MB). Kích thước tối đa: {max_size}MB"
                elif file_size_mb > max_size:  # Soft limit with warning
                    validation_warnings.append(f"File lớn ({file_size_mb:.1f}MB), có thể xử lý chậm")
                    
            except (OSError, PermissionError) as e:
                logger.warning(f"File size check failed: {e}")
                validation_warnings.append("Không thể kiểm tra kích thước file")
            
            # Enhanced security validation with full file read
            try:
                with open(file_path, 'rb') as f:
                    file_data = f.read()  # Read full file for comprehensive validation
                
                is_valid, error_msg = self.security_validator.validate_file(file_data, file_path.name)
                
                # Check if validator has warnings (for lenient mode)
                if hasattr(self.security_validator, 'validation_warnings') and self.security_validator.validation_warnings:
                    validation_warnings.extend(self.security_validator.validation_warnings)
                
                if not is_valid:
                    logger.error(f"Security validation failed for {file_path.name}: {error_msg}")
                    # In lenient mode, try to provide helpful error context
                    if "fallback mode" in str(error_msg):
                        validation_warnings.append("File passed basic validation (compatibility mode)")
                    else:
                        return False, f"File validation failed: {error_msg}"
                else:
                    logger.info(f"Security validation passed for {file_path.name}")
                    
            except (OSError, PermissionError) as e:
                logger.error(f"File read error: {e}")
                return False, f"Không thể đọc file: {str(e)}"
            except Exception as e:
                logger.error(f"Security validation error: {e}")
                return False, f"Lỗi kiểm tra bảo mật: {str(e)}"
            
            # Basic structure validation with fallback
            excel_validation_passed = False
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                if not wb.worksheets:
                    wb.close()
                    validation_warnings.append("File Excel có vấn đề về cấu trúc")
                else:
                    excel_validation_passed = True
                    wb.close()
                    logger.info(f"Excel structure validation passed for {file_path.name}")
                    
            except ImportError:
                logger.warning("openpyxl not available for Excel validation")
                validation_warnings.append("Không thể kiểm tra cấu trúc Excel chi tiết")
                excel_validation_passed = True  # Allow in degraded mode
            except Exception as excel_error:
                logger.warning(f"Excel structure validation failed: {excel_error}")
                validation_warnings.append(f"Cấu trúc Excel có vấn đề: {str(excel_error)}")
                # Don't fail completely - might still be processable
            
            # Compile validation result
            if validation_warnings:
                warning_msg = f"File hợp lệ với {len(validation_warnings)} cảnh báo: {'; '.join(validation_warnings)}"
                logger.warning(warning_msg)
                return True, warning_msg
            else:
                return True, "File hợp lệ"
            
        except SecurityError as se:
            logger.error(f"Security error during validation: {se}")
            return False, f"Lỗi bảo mật: {str(se)}"
        except Exception as e:
            logger.error(f"Unexpected validation error: {e}", exc_info=True)
            return False, f"Lỗi validate file: {str(e)}"
    
    def get_processing_stats(self) -> Dict[str, Any]:
        """Get current processing statistics from secure session state"""
        try:
            # Get stats from secure session state if available
            session_stats = safe_get_session_value('processing_stats', {})
            if session_stats:
                return session_stats.copy()
            return self.processing_stats.copy()
        except Exception as e:
            logger.warning(f"Failed to get processing stats from session: {e}")
            return self.processing_stats.copy()