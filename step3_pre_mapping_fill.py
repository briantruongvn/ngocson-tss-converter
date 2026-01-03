#!/usr/bin/env python3
"""
Step 3: Pre-Mapping Data Fill
Fills empty cells in source sheets before data mapping to ensure data integrity within each sheet.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, Dict, List
import argparse
import sys
import shutil

from common.validation import FileValidator
from common.exceptions import TSConverterError
from common.config import get_config, get_clean_basename

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PreMappingFiller:
    """
    Pre-Mapping Data Filler for Step 2.5
    
    Fills empty cells in source sheets (F, M, C, P types) before mapping:
    - Identifies sheet types based on naming patterns
    - Fills relevant columns within each sheet boundary
    - Preserves data integrity per sheet
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        self.config = get_config()
        
        # Define columns to fill for each sheet type (Material information only)
        self.fill_columns = {
            'M': ['J', 'K', 'L'],  # M-type: Material designation, supplier, material code
            'C': ['I', 'J', 'K'],  # C-type: Material distributor, designation, supplier  
            'P': ['J', 'K', 'L']   # P-type: Material designation, supplier, material code
            # F-type: No filling (skip processing)
        }
    
    def get_sheet_type(self, sheet_name: str) -> Optional[str]:
        """
        Determine the type of sheet based on its name
        
        Args:
            sheet_name: Name of the worksheet
            
        Returns:
            Sheet type ('F', 'M', 'C', 'P') or None if not recognized
        """
        if sheet_name.upper().startswith('F-'):
            return 'F'
        elif sheet_name.upper().startswith('M-'):
            return 'M'
        elif sheet_name.upper().startswith('C-'):
            return 'C'
        elif sheet_name.upper().startswith('P'):
            return 'P'
        else:
            return None
    
    def find_header_row(self, worksheet, header_text: str = "product combination") -> Optional[int]:
        """
        Find row containing the header text (case insensitive)
        
        Args:
            worksheet: openpyxl worksheet object
            header_text: Text to search for in headers
            
        Returns:
            Row number (1-based) or None if not found
        """
        for row in range(1, min(worksheet.max_row + 1, 50)):  # Search first 50 rows
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                cell_value = self.safe_cell_value(cell)
                if cell_value:
                    if header_text.lower() in cell_value.lower():
                        logger.info(f"Found '{header_text}' at row {row}, column {get_column_letter(col)}")
                        return row
        
        logger.warning(f"Header '{header_text}' not found in worksheet")
        return None
    
    def safe_cell_value(self, cell) -> str:
        """
        Safely extract cell value, handling formula errors and edge cases
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Safe string value or empty string if error
        """
        try:
            if cell.value is None:
                return ""
            
            # Check for Excel formula errors
            if isinstance(cell.value, str):
                formula_errors = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!', '#ERROR!']
                if any(error in str(cell.value) for error in formula_errors):
                    logger.warning(f"Formula error detected in {cell.coordinate}: {cell.value} - using empty value")
                    return ""
            
            # Handle numeric values
            if isinstance(cell.value, (int, float)):
                return str(cell.value)
            
            # Handle datetime values  
            from datetime import datetime
            if isinstance(cell.value, datetime):
                return cell.value.strftime('%Y-%m-%d %H:%M:%S')
            
            # Convert to string and clean
            return str(cell.value).strip()
        
        except Exception as e:
            logger.warning(f"Error reading cell {getattr(cell, 'coordinate', 'unknown')}: {e} - using empty value")
            return ""
    
    def find_last_data_row(self, worksheet, start_row: int) -> int:
        """
        Find the last row that contains data starting from start_row
        
        Args:
            worksheet: openpyxl worksheet object
            start_row: Row to start searching from
            
        Returns:
            Last row number (1-based) with data
        """
        last_row = worksheet.max_row
        
        # Search backwards from max_row to find actual last row with data
        for row in range(last_row, start_row - 1, -1):
            has_data = False
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and (not isinstance(cell_value, str) or cell_value.strip()):
                    has_data = True
                    break
            
            if has_data:
                logger.debug(f"Found last data row: {row}")
                return row
        
        logger.debug(f"No data found after row {start_row}")
        return start_row
    
    def fill_column_in_sheet(self, worksheet, column_letter: str, start_row: int, end_row: int) -> int:
        """
        Fill empty cells in a column with data from the last non-empty row in same column
        
        Args:
            worksheet: openpyxl worksheet object
            column_letter: Column letter to fill
            start_row: Starting row (1-based)
            end_row: Ending row (1-based)
            
        Returns:
            Number of cells filled
        """
        col_num = openpyxl.utils.column_index_from_string(column_letter)
        filled_count = 0
        last_non_empty_value = ""
        
        logger.debug(f"Filling column {column_letter} from row {start_row} to {end_row}")
        
        # Process each row from start to end
        for row in range(start_row, end_row + 1):
            current_cell = worksheet.cell(row, col_num)
            current_value = self.safe_cell_value(current_cell)
            
            # Check if current cell has data
            is_empty = (not current_value or current_value.strip() == "")
            
            if not is_empty:
                # Update reference value for subsequent empty cells
                last_non_empty_value = current_value
                logger.debug(f"Found data in {column_letter}{row}: '{current_value}' - will use as reference")
            elif is_empty and last_non_empty_value:
                # Check if cell is merged (can't write to merged cells)
                try:
                    # Fill current empty cell with last non-empty value
                    current_cell.value = last_non_empty_value
                    filled_count += 1
                    logger.debug(f"Filled {column_letter}{row} with '{last_non_empty_value}' from last non-empty")
                except AttributeError as e:
                    # Skip merged cells (they are read-only)
                    logger.debug(f"Skipping merged cell {column_letter}{row}: {e}")
                    continue
        
        return filled_count
    
    def fill_sheet_data(self, sheet_name: str, worksheet) -> Dict[str, int]:
        """
        Fill data in a specific sheet based on its type
        
        Args:
            sheet_name: Name of the worksheet
            worksheet: openpyxl worksheet object
            
        Returns:
            Dictionary with fill counts for each column
        """
        sheet_type = self.get_sheet_type(sheet_name)
        if not sheet_type:
            logger.info(f"Skipping sheet '{sheet_name}' - not a recognized type")
            return {}
        
        # Skip F-type sheets only (process M-type, C-type, and P-type)
        if sheet_type in ['F']:
            logger.info(f"Skipping {sheet_type}-type sheet: {sheet_name} (no material filling needed)")
            return {}
        
        logger.info(f"Processing {sheet_type}-type sheet: {sheet_name}")
        
        # Find header row
        header_row = self.find_header_row(worksheet, "product combination")
        if header_row is None:
            logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
            return {}
        
        # Data starts at header_row + 2
        data_start_row = header_row + 2
        last_data_row = self.find_last_data_row(worksheet, data_start_row)
        
        if last_data_row < data_start_row:
            logger.warning(f"No data to process in {sheet_name}")
            return {}
        
        # Get columns to fill for this sheet type
        columns_to_fill = self.fill_columns.get(sheet_type, [])
        if not columns_to_fill:
            logger.info(f"No columns configured for filling in {sheet_type}-type sheets")
            return {}
        
        # Fill each column
        fill_results = {}
        total_filled = 0
        
        for column_letter in columns_to_fill:
            filled_count = self.fill_column_in_sheet(worksheet, column_letter, data_start_row, last_data_row)
            fill_results[column_letter] = filled_count
            total_filled += filled_count
        
        logger.info(f"Filled {total_filled} cells in {sheet_name} ({sheet_type}-type)")
        if fill_results:
            logger.info(f"  Column breakdown: {fill_results}")
        
        return fill_results
    
    def process_file(self, input_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process input file and fill data in each sheet before mapping
        
        Args:
            input_file: Input Excel file path
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 3: Pre-Mapping Data Fill")
        
        # Validate input file
        try:
            input_path = FileValidator.validate_file_format(input_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = get_clean_basename(input_path)
            output_file = self.output_dir / f"{base_name} - Step3.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Input: {input_path}")
        logger.info(f"Output: {output_file}")
        
        # Copy input file as starting point
        shutil.copy2(str(input_path), str(output_file))
        logger.info("Copied input file as base")
        
        # Load workbook
        workbook = openpyxl.load_workbook(str(output_file))
        
        # Process each sheet
        total_sheets_processed = 0
        total_cells_filled = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Fill data in this sheet
            fill_results = self.fill_sheet_data(sheet_name, worksheet)
            
            if fill_results:
                total_sheets_processed += 1
                sheet_total = sum(fill_results.values())
                total_cells_filled += sheet_total
        
        # Save output file
        try:
            workbook.save(str(output_file))
            logger.info(f"✅ Step 3 completed: {output_file}")
            logger.info(f"📊 Summary: Processed {total_sheets_processed} sheets, filled {total_cells_filled} cells")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        workbook.close()
        
        return str(output_file)

def main():
    """Command line interface for pre-mapping data fill"""
    parser = argparse.ArgumentParser(description='Pre-Mapping Data Fill Step 2.5 - Fill Source Sheets')
    parser.add_argument('input_file', help='Input Excel file (source data)')
    parser.add_argument('-o', '--output', help='Output file path (optional, auto-generates - Step3.xlsx)')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize filler
    filler = PreMappingFiller(args.base_dir)
    
    try:
        result = filler.process_file(args.input_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()