#!/usr/bin/env python3
"""
Step 4: Data Mapping and Transfer
Maps data from source Excel file to Step2 template with specific column mappings.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple, Dict
import argparse
import sys
import shutil

from common.validation import validate_step3_input, FileValidator
from common.exceptions import TSConverterError
from common.config import get_config

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataMapper:
    """
    Data Mapper for Step 3
    
    Maps data from source Excel file to Step2 template based on sheet naming patterns:
    - F-type sheets (F-[Finished products]): Special mapping with header search
    - M-type sheets (M-[Material type]): Direct mapping, continues after existing data
    - C-type sheets (C-[Component type]): Direct mapping, continues after existing data
    - Combines specified columns with delimiter
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        self.config = get_config()
        
        # Column mappings for different sheet types (UPDATED)
        self.f_type_mapping = self.config.get('step3.f_type_mapping', {
            'B': 'Q'
        })
        
        self.m_type_mapping = self.config.get('step3.m_type_mapping', {
            'B': 'Q', 'C': 'B', 'D': 'C', 'J': 'D', 'L': 'E', 'K': 'F',
            'X': 'H', 'Q': 'J', 'R': 'K', 'S': 'L', 'T': 'M', 'U': 'N', 
            'W': 'O', 'AA': 'P'
        })
        
        self.c_type_mapping = self.config.get('step3.c_type_mapping', {
            'B': 'Q', 'C': 'B', 'D': 'C', 'I': 'D', 'J': 'F', 'K': 'E',
            'Q': 'K', 'R': 'L', 'S': 'M', 'T': 'N', 'V': 'O', 'W': 'H', 'Z': 'P'
        })
        
        self.p_type_mapping = self.config.get('step3.p_type_mapping', {
            'B': 'Q', 'C': 'B', 'D': 'C', 'F': 'G', 'J': 'D', 'K': 'F', 'L': 'E',
            'Q': 'J', 'R': 'K', 'S': 'L', 'T': 'M', 'U': 'N', 'W': 'O', 'X': 'H'
        })
    
    def get_sheet_type(self, sheet_name: str) -> Optional[str]:
        """
        Determine the type of sheet based on its name
        
        Args:
            sheet_name: Name of the worksheet
            
        Returns:
            Sheet type ('F', 'M', 'C', 'P') or None if not recognized
        """
        if sheet_name.upper().startswith('F-'):
            logger.info(f"Detected F-type sheet: '{sheet_name}'")
            return 'F'
        elif sheet_name.upper().startswith('M-'):
            logger.info(f"Detected M-type sheet: '{sheet_name}'")
            return 'M'
        elif sheet_name.upper().startswith('C-'):
            logger.info(f"Detected C-type sheet: '{sheet_name}'")
            return 'C'
        elif sheet_name.upper().startswith('P'):
            logger.info(f"Detected P-type sheet: '{sheet_name}'")
            return 'P'
        else:
            logger.debug(f"Sheet '{sheet_name}' does not match F/M/C/P pattern")
            return None
            
    def is_sheet_relevant(self, sheet_name: str, worksheet) -> bool:
        """
        Check if sheet is relevant for processing based on naming pattern
        
        Args:
            sheet_name: Name of the worksheet
            worksheet: openpyxl worksheet object
            
        Returns:
            True if sheet should be processed
        """
        # Check if sheet is empty first
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            cell_value = worksheet.cell(1, 1).value
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                logger.debug(f"Skipping empty sheet '{sheet_name}'")
                return False
        
        # Only process sheets that match F/M/C/P pattern
        sheet_type = self.get_sheet_type(sheet_name)
        if sheet_type is not None:
            logger.info(f"Processing {sheet_type}-type sheet: '{sheet_name}'")
            return True
        
        logger.debug(f"Skipping sheet '{sheet_name}' - not F/M/C/P type")
        return False
    
    def find_header_row(self, worksheet, header_text: str = "product combination") -> Optional[int]:
        """
        Find row containing the header text (case insensitive)
        
        Args:
            worksheet: openpyxl worksheet object
            header_text: Text to search for in headers
            
        Returns:
            Row number (1-based) or None if not found
        """
        for row in range(1, min(worksheet.max_row + 1, 50)):  # Search first 50 rows
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                cell_value = self.safe_cell_value(cell)
                if cell_value:
                    if header_text.lower() in cell_value.lower():
                        logger.info(f"Found '{header_text}' at row {row}, column {get_column_letter(col)}")
                        return row
        
        logger.warning(f"Header '{header_text}' not found in worksheet")
        return None
    
    def get_merged_cell_value(self, worksheet, row: int, col: int) -> str:
        """
        Get cell value with merged cell detection and handling
        
        Args:
            worksheet: openpyxl worksheet object
            row: Row number (1-based)
            col: Column number (1-based)
            
        Returns:
            Cell value as string, handling merged cells appropriately
        """
        try:
            # Check if cell is part of a merged range
            for merged_range in worksheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and 
                    merged_range.min_col <= col <= merged_range.max_col):
                    # Cell is part of a merged range - get value from top-left cell
                    top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                    logger.debug(f"Cell ({row},{col}) is in merged range {merged_range}, extracting from top-left ({merged_range.min_row},{merged_range.min_col})")
                    
                    # Use safe_cell_value to extract the actual value
                    return self.safe_cell_value(top_left_cell)
            
            # Not a merged cell, read directly
            cell = worksheet.cell(row, col)
            return self.safe_cell_value(cell)
            
        except Exception as e:
            logger.warning(f"Error getting merged cell value at ({row},{col}): {e} - using empty value")
            return ""

    def safe_cell_value(self, cell) -> str:
        """
        Safely extract cell value, handling formula errors and edge cases
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Safe string value or empty string if error
        """
        try:
            if cell.value is None:
                return ""
            
            # Check for Excel formula errors
            if isinstance(cell.value, str):
                formula_errors = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!', '#ERROR!']
                if any(error in str(cell.value) for error in formula_errors):
                    logger.warning(f"Formula error detected in {cell.coordinate}: {cell.value} - using empty value")
                    return ""
            
            # Handle numeric values
            if isinstance(cell.value, (int, float)):
                return str(cell.value)
            
            # Handle datetime values  
            from datetime import datetime
            if isinstance(cell.value, datetime):
                return cell.value.strftime('%Y-%m-%d %H:%M:%S')
            
            # Convert to string and clean
            return str(cell.value).strip()
        
        except Exception as e:
            logger.warning(f"Error reading cell {getattr(cell, 'coordinate', 'unknown')}: {e} - using empty value")
            return ""

    def set_column_a_prefix(self, target_ws, target_row: int, sheet_type: str) -> None:
        """
        Set column A prefix based on sheet type
        
        Args:
            target_ws: Target worksheet
            target_row: Target row number
            sheet_type: Sheet type ('F', 'M', 'C', 'P')
        """
        if sheet_type == 'F':
            target_ws.cell(target_row, 1, "Art")
        # M-type, C-type, and P-type: do not set Column A prefix (leave unchanged)

    def handle_f_type_combinations(self, source_ws, source_row: int, target_ws, target_row: int) -> None:
        """
        Handle special F-type combinations: K & L → I
        
        Args:
            source_ws: Source worksheet
            source_row: Source row number
            target_ws: Target worksheet  
            target_row: Target row number
        """
        try:
            k_col = openpyxl.utils.column_index_from_string('K')
            l_col = openpyxl.utils.column_index_from_string('L')
            
            k_value = self.get_merged_cell_value(source_ws, source_row, k_col)
            l_value = self.get_merged_cell_value(source_ws, source_row, l_col)
            
            if k_value and l_value:
                combined = f"{k_value}-{l_value}"
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), combined)
                logger.debug(f"F-type combination K&L→I: {k_value}-{l_value} = {combined}")
            elif k_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), k_value)
                logger.debug(f"F-type combination K→I: {k_value}")
            elif l_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), l_value)
                logger.debug(f"F-type combination L→I: {l_value}")
        except Exception as e:
            logger.warning(f"Error handling F-type combinations at row {source_row}: {e}")

    def handle_m_type_combinations(self, source_ws, source_row: int, target_ws, target_row: int) -> None:
        """
        Handle special M-type combinations: O & P → I
        
        Args:
            source_ws: Source worksheet
            source_row: Source row number
            target_ws: Target worksheet  
            target_row: Target row number
        """
        try:
            o_col = openpyxl.utils.column_index_from_string('O')
            p_col = openpyxl.utils.column_index_from_string('P')
            
            o_value = self.get_merged_cell_value(source_ws, source_row, o_col)
            p_value = self.get_merged_cell_value(source_ws, source_row, p_col)
            
            if o_value and p_value:
                combined = f"{o_value}-{p_value}"
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), combined)
                logger.debug(f"M-type combination O&P→I: {o_value}-{p_value} = {combined}")
            elif o_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), o_value)
            elif p_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), p_value)
        except Exception as e:
            logger.warning(f"Error handling M-type combinations at row {source_row}: {e}")

    def handle_c_type_combinations(self, source_ws, source_row: int, target_ws, target_row: int) -> None:
        """
        Handle special C-type combinations: N & O → I
        
        Args:
            source_ws: Source worksheet
            source_row: Source row number
            target_ws: Target worksheet
            target_row: Target row number
        """
        try:
            n_col = openpyxl.utils.column_index_from_string('N')
            o_col = openpyxl.utils.column_index_from_string('O')
            
            n_value = self.get_merged_cell_value(source_ws, source_row, n_col)
            o_value = self.get_merged_cell_value(source_ws, source_row, o_col)
            
            if n_value and o_value:
                combined = f"{n_value}-{o_value}"
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), combined)
                logger.debug(f"C-type combination N&O→I: {n_value}-{o_value} = {combined}")
            elif n_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), n_value)
            elif o_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), o_value)
        except Exception as e:
            logger.warning(f"Error handling C-type combinations at row {source_row}: {e}")

    def handle_p_type_combinations(self, source_ws, source_row: int, target_ws, target_row: int) -> None:
        """
        Handle special P-type combinations: O & P → I
        
        Args:
            source_ws: Source worksheet
            source_row: Source row number
            target_ws: Target worksheet  
            target_row: Target row number
        """
        try:
            o_col = openpyxl.utils.column_index_from_string('O')
            p_col = openpyxl.utils.column_index_from_string('P')
            
            o_value = self.get_merged_cell_value(source_ws, source_row, o_col)
            p_value = self.get_merged_cell_value(source_ws, source_row, p_col)
            
            if o_value and p_value:
                combined = f"{o_value}-{p_value}"
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), combined)
                logger.debug(f"P-type combination O&P→I: {o_value}-{p_value} = {combined}")
            elif o_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), o_value)
                logger.debug(f"P-type combination O→I: {o_value}")
            elif p_value:
                target_ws.cell(target_row, openpyxl.utils.column_index_from_string('I'), p_value)
                logger.debug(f"P-type combination P→I: {p_value}")
        except Exception as e:
            logger.warning(f"Error handling P-type combinations at row {source_row}: {e}")
    
    def combine_columns(self, worksheet, row: int, col1: str, col2: str, delimiter: str = "-") -> str:
        """
        Combine values from two columns with delimiter using merged cell aware reading
        
        Args:
            worksheet: openpyxl worksheet object
            row: Row number (1-based)
            col1: First column letter
            col2: Second column letter
            delimiter: Delimiter between values
            
        Returns:
            Combined string value
        """
        col1_num = openpyxl.utils.column_index_from_string(col1)
        col2_num = openpyxl.utils.column_index_from_string(col2)
        
        # Use merged cell aware reading
        str1 = self.get_merged_cell_value(worksheet, row, col1_num)
        str2 = self.get_merged_cell_value(worksheet, row, col2_num)
        
        # Combine with delimiter
        if str1 and str2:
            return f"{str1}{delimiter}{str2}"
        elif str1:
            return str1
        elif str2:
            return str2
        else:
            return ""
    
    def map_f_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from F-type sheet (F-[Finished products])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping F-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using merged cell aware reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.get_merged_cell_value(source_ws, source_row, col)
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Set column A prefix for F-type
            self.set_column_a_prefix(target_ws, current_target_row, 'F')
            
            # Handle F-type special combination: K & L → I
            self.handle_f_type_combinations(source_ws, source_row, target_ws, current_target_row)
            
            # Apply F-type column mappings using merged cell aware approach
            source_values = {}
            for source_col in self.f_type_mapping.keys():
                try:
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    source_value = self.get_merged_cell_value(source_ws, source_row, source_col_num)
                    if source_value:
                        source_values[source_col] = source_value
                except Exception as e:
                    logger.warning(f"Error reading F-type source column {source_col}: {e}")
                    
            for source_col, target_col in self.f_type_mapping.items():
                if source_col in source_values:
                    try:
                        target_col_num = openpyxl.utils.column_index_from_string(target_col)
                        target_ws.cell(current_target_row, target_col_num, source_values[source_col])
                        logger.debug(f"F-type {source_col} -> {target_col}: '{source_values[source_col][:30]}...'")
                    except Exception as e:
                        logger.warning(f"Error mapping F-type {source_col} -> {target_col}: {e}")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from F-type sheet")
        return current_target_row
    
    def map_m_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from M-type sheet (M-[Material type])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping M-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using merged cell aware reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.get_merged_cell_value(source_ws, source_row, col)
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Set column A prefix for M-type
            self.set_column_a_prefix(target_ws, current_target_row, 'M')
            
            # Handle M-type special combination: O & P → I
            self.handle_m_type_combinations(source_ws, source_row, target_ws, current_target_row)
            
            # Apply M-type column mappings using merged cell aware approach
            # First, collect all source values to avoid overwriting issues
            source_values = {}
            for source_col in self.m_type_mapping.keys():
                try:
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    source_value = self.get_merged_cell_value(source_ws, source_row, source_col_num)
                    if source_value:
                        source_values[source_col] = source_value
                except Exception as e:
                    logger.warning(f"Error reading source column {source_col}: {e}")
                    
            # Then apply all mappings
            for source_col, target_col in self.m_type_mapping.items():
                if source_col in source_values:
                    try:
                        target_col_num = openpyxl.utils.column_index_from_string(target_col)
                        target_ws.cell(current_target_row, target_col_num, source_values[source_col])
                        logger.debug(f"M-type {source_col} -> {target_col}: '{source_values[source_col][:30]}...'")
                    except Exception as e:
                        logger.warning(f"Error mapping {source_col} -> {target_col}: {e}")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from M-type sheet")
        return current_target_row
        
    def map_c_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from C-type sheet (C-[Component type])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping C-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using merged cell aware reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.get_merged_cell_value(source_ws, source_row, col)
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Set column A prefix for C-type (not specified in requirements)
            self.set_column_a_prefix(target_ws, current_target_row, 'C')
            
            # Handle C-type special combination: N & O → I
            self.handle_c_type_combinations(source_ws, source_row, target_ws, current_target_row)
            
            # Apply C-type column mappings using merged cell aware approach
            source_values = {}
            for source_col in self.c_type_mapping.keys():
                try:
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    source_value = self.get_merged_cell_value(source_ws, source_row, source_col_num)
                    if source_value:
                        source_values[source_col] = source_value
                except Exception as e:
                    logger.warning(f"Error reading C-type source column {source_col}: {e}")
                    
            for source_col, target_col in self.c_type_mapping.items():
                if source_col in source_values:
                    try:
                        target_col_num = openpyxl.utils.column_index_from_string(target_col)
                        target_ws.cell(current_target_row, target_col_num, source_values[source_col])
                        logger.debug(f"C-type {source_col} -> {target_col}: '{source_values[source_col][:30]}...'")
                    except Exception as e:
                        logger.warning(f"Error mapping C-type {source_col} -> {target_col}: {e}")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from C-type sheet")
        return current_target_row
        
    def map_p_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from P-type sheet (P-type sheets like PInorganic coating)
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping P-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using merged cell aware reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.get_merged_cell_value(source_ws, source_row, col)
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Set column A prefix for P-type
            self.set_column_a_prefix(target_ws, current_target_row, 'P')
            
            # Handle P-type special combination: O & P → I
            self.handle_p_type_combinations(source_ws, source_row, target_ws, current_target_row)
            
            # Apply P-type column mappings using merged cell aware approach
            source_values = {}
            for source_col in self.p_type_mapping.keys():
                try:
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    source_value = self.get_merged_cell_value(source_ws, source_row, source_col_num)
                    if source_value:
                        source_values[source_col] = source_value
                except Exception as e:
                    logger.warning(f"Error reading P-type source column {source_col}: {e}")
                    
            for source_col, target_col in self.p_type_mapping.items():
                if source_col in source_values:
                    try:
                        target_col_num = openpyxl.utils.column_index_from_string(target_col)
                        target_ws.cell(current_target_row, target_col_num, source_values[source_col])
                        logger.debug(f"P-type {source_col} -> {target_col}: '{source_values[source_col][:30]}...'")
                    except Exception as e:
                        logger.warning(f"Error mapping P-type {source_col} -> {target_col}: {e}")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from P-type sheet")
        return current_target_row
    
    def process_file(self, input_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step 3 output file and map data to Step2 template (auto-detected)
        
        Args:
            input_file: Input Excel file from Step 3 (e.g., Input-3 - Step3.xlsx)
            output_file: Optional output file path (if None, auto-generate Step4)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 4: Data Mapping")
        
        # Validate input file
        try:
            input_path = FileValidator.validate_file_format(input_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Use input file directly (should be Step 3 output)
        base_name = input_path.stem  # e.g., "Input-3 - Step3"
        
        # Extract original filename from Step3 naming
        if " - Step3" in base_name:
            original_name = base_name.replace(" - Step3", "")
        else:
            original_name = base_name
        
        # Auto-detect Step2 template file using original name
        step2_file = self.output_dir / f"{original_name} - Step2.xlsx"
        
        try:
            step2_path = FileValidator.validate_file_format(step2_file)
        except TSConverterError as e:
            logger.error(f"Step2 template not found: {step2_file}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            output_file = self.output_dir / f"{original_name} - Step4.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Input Source: {input_path}")
        logger.info(f"Step2 Template: {step2_path}")
        logger.info(f"Output: {output_file}")
        
        # Copy Step2 file as starting point
        shutil.copy2(str(step2_path), str(output_file))
        logger.info("Copied Step2 template as base")
        
        # Load workbooks
        source_wb = openpyxl.load_workbook(str(input_path))
        target_wb = openpyxl.load_workbook(str(output_file))
        target_ws = target_wb.active
        
        # Find next available row in target (after existing data)
        next_row = 11  # Start from row 11 (after headers and article data)
        while target_ws.cell(next_row, 2).value is not None:  # Check column B
            next_row += 1
        
        logger.info(f"Starting data mapping at target row {next_row}")
        
        # Process each sheet in source file
        for sheet_name in source_wb.sheetnames:
            worksheet = source_wb[sheet_name]
            
            # Check if sheet is relevant for processing
            if not self.is_sheet_relevant(sheet_name, worksheet):
                continue
            
            # Get sheet type and process accordingly
            sheet_type = self.get_sheet_type(sheet_name)
            
            if sheet_type == 'F':
                logger.info(f"Processing F-type sheet: {sheet_name}")
                
                # Find header row for F-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_f_type_data(worksheet, target_ws, data_start_row, next_row)
                
            elif sheet_type == 'M':
                logger.info(f"Processing M-type sheet: {sheet_name}")
                
                # Find header row for M-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_m_type_data(worksheet, target_ws, data_start_row, next_row)
                
            elif sheet_type == 'C':
                logger.info(f"Processing C-type sheet: {sheet_name}")
                
                # Find header row for C-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_c_type_data(worksheet, target_ws, data_start_row, next_row)
                
            elif sheet_type == 'P':
                logger.info(f"Processing P-type sheet: {sheet_name}")
                
                # Find header row for P-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_p_type_data(worksheet, target_ws, data_start_row, next_row)
        
        # Save output file
        try:
            target_wb.save(str(output_file))
            logger.info(f"✅ Step 4 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        source_wb.close()
        target_wb.close()
        
        return str(output_file)

def main():
    """Command line interface for data mapping"""
    parser = argparse.ArgumentParser(description='Data Mapper Step 3 - Map Excel Data')
    parser.add_argument('input_file', help='Input Excel file to extract data from (e.g., input-1.xlsx)')
    parser.add_argument('-o', '--output', help='Output file path (optional, auto-generates Step4.xlsx)')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize mapper
    mapper = DataMapper(args.base_dir)
    
    try:
        # Process file directly (should be Step 3 output)
        result = mapper.process_file(args.input_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()