#!/usr/bin/env python3
"""
Step 3: Data Mapping and Transfer
Maps data from source Excel file to Step2 template with specific column mappings.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple, Dict
import argparse
import sys
import shutil
from copy import copy

from common.validation import validate_step3_input, FileValidator
from common.exceptions import TSConverterError
from common.config import get_config

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataMapper:
    """
    Data Mapper for Step 3
    
    Maps data from source Excel file to Step2 template based on sheet naming patterns:
    - F-type sheets (F-[Finished products]): Special mapping with header search
    - M-type sheets (M-[Material type]): Direct mapping, continues after existing data
    - C-type sheets (C-[Component type]): Direct mapping, continues after existing data
    - Combines specified columns with delimiter
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        self.config = get_config()
        
        # Column mappings for different sheet types
        self.f_type_mapping = self.config.get('step3.f_type_mapping', {
            'C': 'D', 'H': 'F', 'KL': 'I', 'M': 'J', 'N': 'K', 
            'O': 'L', 'P': 'M', 'Q': 'N', 'S': 'O', 'T': 'H', 'W': 'P'
        })
        
        self.m_type_mapping = self.config.get('step3.m_type_mapping', {
            'B': 'B', 'C': 'C', 'I': 'D', 'J': 'F', 'K': 'E', 
            'NO': 'I', 'P': 'J', 'Q': 'K', 'R': 'L', 'S': 'M', 'T': 'N', 'W': 'H', 'Z': 'P'
        })
        
        self.c_type_mapping = self.config.get('step3.c_type_mapping', {
            'B': 'B', 'C': 'C', 'H': 'D', 'I': 'F', 'J': 'E', 
            'MN': 'I', 'O': 'J', 'P': 'K', 'Q': 'L', 'R': 'M', 'S': 'N', 'V': 'H', 'Y': 'P'
        })
    
    def get_sheet_type(self, sheet_name: str) -> Optional[str]:
        """
        Determine the type of sheet based on its name
        
        Args:
            sheet_name: Name of the worksheet
            
        Returns:
            Sheet type ('F', 'M', 'C') or None if not recognized
        """
        if sheet_name.upper().startswith('F-'):
            logger.info(f"Detected F-type sheet: '{sheet_name}'")
            return 'F'
        elif sheet_name.upper().startswith('M-'):
            logger.info(f"Detected M-type sheet: '{sheet_name}'")
            return 'M'
        elif sheet_name.upper().startswith('C-'):
            logger.info(f"Detected C-type sheet: '{sheet_name}'")
            return 'C'
        else:
            logger.debug(f"Sheet '{sheet_name}' does not match F/M/C pattern")
            return None
            
    def is_sheet_relevant(self, sheet_name: str, worksheet) -> bool:
        """
        Check if sheet is relevant for processing based on naming pattern
        
        Args:
            sheet_name: Name of the worksheet
            worksheet: openpyxl worksheet object
            
        Returns:
            True if sheet should be processed
        """
        # Check if sheet is empty first
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            cell_value = worksheet.cell(1, 1).value
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                logger.debug(f"Skipping empty sheet '{sheet_name}'")
                return False
        
        # Only process sheets that match F/M/C pattern
        sheet_type = self.get_sheet_type(sheet_name)
        if sheet_type is not None:
            logger.info(f"Processing {sheet_type}-type sheet: '{sheet_name}'")
            return True
        
        logger.debug(f"Skipping sheet '{sheet_name}' - not F/M/C type")
        return False
    
    def find_header_row(self, worksheet, header_text: str = "product combination") -> Optional[int]:
        """
        Find row containing the header text (case insensitive)
        
        Args:
            worksheet: openpyxl worksheet object
            header_text: Text to search for in headers
            
        Returns:
            Row number (1-based) or None if not found
        """
        for row in range(1, min(worksheet.max_row + 1, 20)):  # Optimized: search first 20 rows only
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                cell_value = self.safe_cell_value(cell)
                if cell_value:
                    if header_text.lower() in cell_value.lower():
                        logger.info(f"Found '{header_text}' at row {row}, column {get_column_letter(col)}")
                        return row
        
        logger.warning(f"Header '{header_text}' not found in worksheet")
        return None
    
    def safe_cell_value(self, cell) -> str:
        """
        Safely extract cell value, handling formula errors and edge cases
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Safe string value or empty string if error
        """
        try:
            if cell.value is None:
                return ""
            
            # Check for Excel formula errors
            if isinstance(cell.value, str):
                formula_errors = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!', '#ERROR!']
                if any(error in str(cell.value) for error in formula_errors):
                    logger.warning(f"Formula error detected in {cell.coordinate}: {cell.value} - using empty value")
                    return ""
            
            # Handle numeric values
            if isinstance(cell.value, (int, float)):
                return str(cell.value)
            
            # Handle datetime values  
            from datetime import datetime
            if isinstance(cell.value, datetime):
                return cell.value.strftime('%Y-%m-%d %H:%M:%S')
            
            # Convert to string and clean
            return str(cell.value).strip()
        
        except Exception as e:
            logger.warning(f"Error reading cell {getattr(cell, 'coordinate', 'unknown')}: {e} - using empty value")
            return ""
    
    def combine_columns(self, worksheet, row: int, col1: str, col2: str, delimiter: str = "-") -> str:
        """
        Combine values from two columns with delimiter using safe cell reading
        
        Args:
            worksheet: openpyxl worksheet object
            row: Row number (1-based)
            col1: First column letter
            col2: Second column letter
            delimiter: Delimiter between values
            
        Returns:
            Combined string value
        """
        col1_num = openpyxl.utils.column_index_from_string(col1)
        col2_num = openpyxl.utils.column_index_from_string(col2)
        
        cell1 = worksheet.cell(row, col1_num)
        cell2 = worksheet.cell(row, col2_num)
        
        # Use safe cell reading to handle formula errors
        str1 = self.safe_cell_value(cell1)
        str2 = self.safe_cell_value(cell2)
        
        # Combine with delimiter
        if str1 and str2:
            return f"{str1}{delimiter}{str2}"
        elif str1:
            return str1
        elif str2:
            return str2
        else:
            return ""
    
    def map_f_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from F-type sheet (F-[Finished products])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping F-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using safe cell reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.safe_cell_value(source_ws.cell(source_row, col))
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Apply column mappings
            for source_col, target_col in self.f_type_mapping.items():
                if source_col == 'KL':  # Special case: combine K & L
                    combined_value = self.combine_columns(source_ws, source_row, 'K', 'L')
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    target_ws.cell(current_target_row, target_col_num, combined_value)
                    logger.debug(f"Combined K&L -> {target_col}: '{combined_value}'")
                else:
                    # Single column mapping
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    
                    source_cell = source_ws.cell(source_row, source_col_num)
                    source_value = self.safe_cell_value(source_cell)
                    if source_value:
                        target_ws.cell(current_target_row, target_col_num, source_value)
                        logger.debug(f"{source_col} -> {target_col}: '{source_value}'")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from F-type sheet")
        return current_target_row
    
    def map_m_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from M-type sheet (M-[Material type])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping M-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using safe cell reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.safe_cell_value(source_ws.cell(source_row, col))
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Apply column mappings
            for source_col, target_col in self.m_type_mapping.items():
                if source_col == 'NO':  # Special case: combine N & O
                    combined_value = self.combine_columns(source_ws, source_row, 'N', 'O')
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    target_ws.cell(current_target_row, target_col_num, combined_value)
                    logger.debug(f"Combined N&O -> {target_col}: '{combined_value}'")
                else:
                    # Single column mapping
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    
                    source_cell = source_ws.cell(source_row, source_col_num)
                    source_value = self.safe_cell_value(source_cell)
                    if source_value:
                        target_ws.cell(current_target_row, target_col_num, source_value)
                        logger.debug(f"{source_col} -> {target_col}: '{source_value}'")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from M-type sheet")
        return current_target_row
        
    def map_c_type_data(self, source_ws, target_ws, start_row: int, target_start_row: int) -> int:
        """
        Map data from C-type sheet (C-[Component type])
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in source (1-based)
            target_start_row: Starting row in target (1-based)
            
        Returns:
            Next available row in target worksheet
        """
        logger.info(f"Mapping C-type data from row {start_row}")
        current_target_row = target_start_row
        
        # Process each row until empty
        for source_row in range(start_row, source_ws.max_row + 1):
            # Check if row has any data using safe cell reading
            has_data = False
            for col in range(1, source_ws.max_column + 1):
                cell_value = self.safe_cell_value(source_ws.cell(source_row, col))
                if cell_value:
                    has_data = True
                    break
            
            if not has_data:
                logger.debug(f"Stopping at empty row {source_row}")
                break
            
            logger.debug(f"Processing source row {source_row} -> target row {current_target_row}")
            
            # Apply column mappings
            for source_col, target_col in self.c_type_mapping.items():
                if source_col == 'MN':  # Special case: combine M & N
                    combined_value = self.combine_columns(source_ws, source_row, 'M', 'N')
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    target_ws.cell(current_target_row, target_col_num, combined_value)
                    logger.debug(f"Combined M&N -> {target_col}: '{combined_value}'")
                else:
                    # Single column mapping
                    source_col_num = openpyxl.utils.column_index_from_string(source_col)
                    target_col_num = openpyxl.utils.column_index_from_string(target_col)
                    
                    source_cell = source_ws.cell(source_row, source_col_num)
                    source_value = self.safe_cell_value(source_cell)
                    if source_value:
                        target_ws.cell(current_target_row, target_col_num, source_value)
                        logger.debug(f"{source_col} -> {target_col}: '{source_value}'")
            
            current_target_row += 1
        
        rows_mapped = current_target_row - target_start_row
        logger.info(f"Mapped {rows_mapped} rows from C-type sheet")
        return current_target_row
    
    def process_file(self, source_file: Union[str, Path],
                    step2_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process source file and map data to Step2 template
        
        Args:
            source_file: Source Excel file to extract data from
            step2_file: Step2 template file
            output_file: Optional output file path (if None, auto-generate Step3)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 3: Data Mapping")
        
        # Validate input files
        try:
            source_path = FileValidator.validate_file_format(source_file)
            # Skip Step3 validation as Step2 only has 3 template rows, no data rows
            # validate_step3_input(step2_file)
            step2_path = Path(step2_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step2_path.stem.replace(" - Step2", "")
            output_file = self.output_dir / f"{base_name} - Step3.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Source: {source_path}")
        logger.info(f"Step2 Template: {step2_path}")
        logger.info(f"Output: {output_file}")
        
        # OPTIMIZATION: Load Step2 template directly instead of copying file
        logger.info("Loading Step2 template directly (no file copy)")
        
        # Load workbooks (source as read_only for performance, step2 for copying structure)
        source_wb = openpyxl.load_workbook(str(source_path), read_only=True)
        step2_wb = openpyxl.load_workbook(str(step2_path), read_only=True)
        
        # Create new workbook with Step2 structure (in-memory copy)
        target_wb = openpyxl.Workbook()
        target_ws = target_wb.active
        step2_ws = step2_wb.active
        
        # Copy Step2 data to new workbook (in-memory, much faster)
        for row_idx, row in enumerate(step2_ws.iter_rows(values_only=True), 1):
            for col_idx, value in enumerate(row, 1):
                if value is not None:
                    target_ws.cell(row_idx, col_idx, value)
        
        # Copy formatting from Step2 (basic cell styles)
        for row_idx in range(1, min(step2_ws.max_row + 1, 10)):  # Only copy first 10 rows formatting
            for col_idx in range(1, step2_ws.max_column + 1):
                source_cell = step2_ws.cell(row_idx, col_idx)
                target_cell = target_ws.cell(row_idx, col_idx)
                if hasattr(source_cell, 'font'):
                    target_cell.font = copy(source_cell.font)
                if hasattr(source_cell, 'fill'):
                    target_cell.fill = copy(source_cell.fill)
        
        step2_wb.close()  # Close source workbook to free memory
        target_ws = target_wb.active
        
        # Find next available row in target (after existing data)
        next_row = 4  # Start from row 4 (after headers and article data)
        while target_ws.cell(next_row, 2).value is not None:  # Check column B
            next_row += 1
        
        logger.info(f"Starting data mapping at target row {next_row}")
        
        # Process each sheet in source file
        for sheet_name in source_wb.sheetnames:
            worksheet = source_wb[sheet_name]
            
            # Check if sheet is relevant for processing
            if not self.is_sheet_relevant(sheet_name, worksheet):
                continue
            
            # Get sheet type and process accordingly
            sheet_type = self.get_sheet_type(sheet_name)
            
            if sheet_type == 'F':
                logger.info(f"Processing F-type sheet: {sheet_name}")
                
                # Find header row for F-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_f_type_data(worksheet, target_ws, data_start_row, next_row)
                
            elif sheet_type == 'M':
                logger.info(f"Processing M-type sheet: {sheet_name}")
                
                # Find header row for M-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_m_type_data(worksheet, target_ws, data_start_row, next_row)
                
            elif sheet_type == 'C':
                logger.info(f"Processing C-type sheet: {sheet_name}")
                
                # Find header row for C-type sheets
                header_row = self.find_header_row(worksheet, "product combination")
                if header_row is None:
                    logger.warning(f"No 'product combination' found in {sheet_name}, skipping")
                    continue
                
                # Data starts at header_row + 2
                data_start_row = header_row + 2
                next_row = self.map_c_type_data(worksheet, target_ws, data_start_row, next_row)
        
        # Save output file
        try:
            target_wb.save(str(output_file))
            logger.info(f"✅ Step 3 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        source_wb.close()
        target_wb.close()
        
        return str(output_file)

def main():
    """Command line interface for data mapping"""
    parser = argparse.ArgumentParser(description='Data Mapper Step 3 - Map Excel Data')
    parser.add_argument('source_file', help='Source Excel file to extract data from')
    parser.add_argument('step2_file', help='Step2 template file (*.xlsx)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize mapper
    mapper = DataMapper(args.base_dir)
    
    try:
        result = mapper.process_file(args.source_file, args.step2_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()