"""
File validation utilities for Excel Template Converter
Provides robust validation for file formats, structure, and data integrity.
"""

import os
from pathlib import Path
from typing import Union, List, Optional, Dict, Any, Tuple
import logging

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

from .exceptions import (
    FileFormatError, FileAccessError, WorksheetNotFoundError, 
    ColumnMissingError, HeaderNotFoundError, InsufficientDataError,
    ValidationError
)

logger = logging.getLogger(__name__)


class FileValidator:
    """
    File validation utilities for Excel files
    Validates format, accessibility, and basic structure
    """
    
    SUPPORTED_EXTENSIONS = ['.xlsx']
    
    @classmethod
    def validate_file_exists(cls, file_path: Union[str, Path]) -> Path:
        """
        Validate that file exists and is accessible
        
        Args:
            file_path: Path to file
            
        Returns:
            Path object if valid
            
        Raises:
            FileAccessError: If file doesn't exist or isn't accessible
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason="File does not exist"
            )
        
        if not path.is_file():
            raise FileAccessError(
                file_path=str(path),
                operation="read", 
                reason="Path is not a file"
            )
        
        if not os.access(path, os.R_OK):
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason="File is not readable"
            )
        
        return path
    
    @classmethod
    def validate_file_format(cls, file_path: Union[str, Path]) -> Path:
        """
        Validate file format is supported Excel format
        
        Args:
            file_path: Path to file
            
        Returns:
            Path object if valid
            
        Raises:
            FileFormatError: If format is not supported
        """
        path = cls.validate_file_exists(file_path)
        
        # Check file extension
        if path.suffix.lower() not in cls.SUPPORTED_EXTENSIONS:
            raise FileFormatError(
                file_path=str(path),
                expected_format=".xlsx",
                actual_format=path.suffix
            )
        
        # File extension validation already done above
        # MIME type check removed for deployment compatibility
        
        # Try to open with openpyxl to validate structure
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping Excel structure validation")
            return path
            
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            wb.close()
        except Exception as e:
            raise FileFormatError(
                file_path=str(path),
                expected_format="Valid Excel file",
                actual_format=f"Corrupted or invalid: {str(e)}"
            )
        
        return path
    
    @classmethod
    def validate_output_writable(cls, file_path: Union[str, Path]) -> Path:
        """
        Validate output file path is writable
        
        Args:
            file_path: Path to output file
            
        Returns:
            Path object if valid
            
        Raises:
            FileAccessError: If path is not writable
        """
        path = Path(file_path)
        
        # Check if parent directory exists and is writable
        parent = path.parent
        if not parent.exists():
            try:
                parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                raise FileAccessError(
                    file_path=str(path),
                    operation="write",
                    reason=f"Cannot create parent directory: {str(e)}"
                )
        
        if not os.access(parent, os.W_OK):
            raise FileAccessError(
                file_path=str(path),
                operation="write",
                reason="Parent directory is not writable"
            )
        
        # If file exists, check if it's writable
        if path.exists():
            if not os.access(path, os.W_OK):
                raise FileAccessError(
                    file_path=str(path),
                    operation="write",
                    reason="File exists but is not writable"
                )
        
        return path


class ExcelStructureValidator:
    """
    Validates Excel file structure and content
    """
    
    @classmethod
    def validate_worksheets_exist(cls, file_path: Union[str, Path], 
                                 required_sheets: Optional[List[str]] = None) -> List[str]:
        """
        Validate that required worksheets exist
        
        Args:
            file_path: Path to Excel file
            required_sheets: List of required sheet names (optional)
            
        Returns:
            List of available sheet names
            
        Raises:
            WorksheetNotFoundError: If required sheets are missing
        """
        path = FileValidator.validate_file_format(file_path)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping worksheet validation")
            return []
        
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            available_sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason=f"Failed to read worksheets: {str(e)}"
            )
        
        if required_sheets:
            missing_sheets = [sheet for sheet in required_sheets if sheet not in available_sheets]
            if missing_sheets:
                raise WorksheetNotFoundError(
                    worksheet_name=", ".join(missing_sheets),
                    available_sheets=available_sheets
                )
        
        return available_sheets
    
    @classmethod
    def validate_columns_exist(cls, file_path: Union[str, Path], 
                             required_columns: List[str],
                             worksheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Validate that required columns exist in worksheet
        
        Args:
            file_path: Path to Excel file
            required_columns: List of required column letters (e.g., ['A', 'B', 'C'])
            worksheet_name: Name of worksheet to check (if None, use active sheet)
            
        Returns:
            Dictionary with validation results
            
        Raises:
            ColumnMissingError: If required columns are missing
        """
        path = FileValidator.validate_file_format(file_path)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping column validation")
            return {"worksheet_name": worksheet_name or "unknown", "available_columns": [], "max_column": 0, "required_columns": required_columns}
        
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            
            if worksheet_name:
                if worksheet_name not in wb.sheetnames:
                    raise WorksheetNotFoundError(worksheet_name, wb.sheetnames)
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                worksheet_name = ws.title
            
            # Check if columns have data or headers
            max_col = ws.max_column
            available_columns = [openpyxl.utils.get_column_letter(i) for i in range(1, max_col + 1)]
            
            missing_columns = [col for col in required_columns if col not in available_columns]
            
            wb.close()
            
            if missing_columns:
                raise ColumnMissingError(
                    missing_columns=missing_columns,
                    worksheet_name=worksheet_name
                )
            
            return {
                "worksheet_name": worksheet_name,
                "available_columns": available_columns,
                "max_column": max_col,
                "required_columns": required_columns
            }
            
        except (ColumnMissingError, WorksheetNotFoundError):
            raise
        except Exception as e:
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason=f"Failed to validate columns: {str(e)}"
            )
    
    @classmethod
    def validate_headers_exist_graceful(cls, file_path: Union[str, Path],
                                       required_headers: List[str],
                                       search_rows: int = 10,
                                       worksheet_name: Optional[str] = None,
                                       graceful: bool = False) -> Tuple[Dict[str, Any], List[str]]:
        """
        Validate that required headers exist in worksheet with optional graceful mode
        
        Args:
            file_path: Path to Excel file
            required_headers: List of required header texts
            search_rows: Number of rows to search for headers
            worksheet_name: Name of worksheet to check
            graceful: If True, return warnings for missing headers instead of raising
            
        Returns:
            Tuple of (result_dict, warnings_list)
            
        Raises:
            HeaderNotFoundError: If required headers are missing and graceful=False
        """
        path = FileValidator.validate_file_format(file_path)
        warnings = []
        
        if not OPENPYXL_AVAILABLE:
            warning = "openpyxl not available, skipping header validation"
            warnings.append(warning)
            logger.warning(warning)
            return {"worksheet_name": worksheet_name or "unknown", "found_headers": {}, "search_rows": search_rows}, warnings
        
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            
            if worksheet_name:
                if worksheet_name not in wb.sheetnames:
                    wb.close()
                    if graceful:
                        warning = f"Worksheet '{worksheet_name}' not found, using active sheet"
                        warnings.append(warning)
                        wb = openpyxl.load_workbook(str(path), read_only=True)
                        ws = wb.active
                        worksheet_name = ws.title
                    else:
                        raise WorksheetNotFoundError(worksheet_name, wb.sheetnames)
                else:
                    ws = wb[worksheet_name]
            else:
                ws = wb.active
                worksheet_name = ws.title
            
            found_headers = {}
            missing_headers = []
            
            # Search for headers in first N rows
            for header in required_headers:
                found = False
                for row in range(1, min(search_rows + 1, ws.max_row + 1)):
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row, col).value
                        if cell_value and isinstance(cell_value, str):
                            if header.lower() in cell_value.lower():
                                found_headers[header] = {"row": row, "column": col, "value": cell_value}
                                found = True
                                break
                    if found:
                        break
                
                if not found:
                    missing_headers.append(header)
            
            wb.close()
            
            if missing_headers:
                warning = f"Missing headers in '{worksheet_name}': {', '.join(missing_headers)}"
                warnings.append(warning)
                
                if not graceful:
                    search_area = f"first {search_rows} rows of '{worksheet_name}'"
                    raise HeaderNotFoundError(
                        header_name=", ".join(missing_headers),
                        search_area=search_area
                    )
            
            return {
                "worksheet_name": worksheet_name,
                "found_headers": found_headers,
                "search_rows": search_rows,
                "missing_headers": missing_headers
            }, warnings
            
        except (HeaderNotFoundError, WorksheetNotFoundError):
            if not graceful:
                raise
            return {"worksheet_name": worksheet_name or "unknown", "found_headers": {}, "search_rows": search_rows}, warnings
        except Exception as e:
            error_msg = f"Failed to validate headers: {str(e)}"
            if graceful:
                warnings.append(error_msg)
                return {"worksheet_name": worksheet_name or "unknown", "found_headers": {}, "search_rows": search_rows}, warnings
            else:
                raise FileAccessError(
                    file_path=str(path),
                    operation="read",
                    reason=error_msg
                )
    
    @classmethod
    def validate_headers_exist(cls, file_path: Union[str, Path],
                              required_headers: List[str],
                              search_rows: int = 10,
                              worksheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Validate that required headers exist in worksheet
        
        Args:
            file_path: Path to Excel file
            required_headers: List of required header texts
            search_rows: Number of rows to search for headers
            worksheet_name: Name of worksheet to check
            
        Returns:
            Dictionary with found headers and their positions
            
        Raises:
            HeaderNotFoundError: If required headers are missing
        """
        path = FileValidator.validate_file_format(file_path)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping header validation")
            return {"worksheet_name": worksheet_name or "unknown", "found_headers": {}, "search_rows": search_rows}
        
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            
            if worksheet_name:
                if worksheet_name not in wb.sheetnames:
                    raise WorksheetNotFoundError(worksheet_name, wb.sheetnames)
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                worksheet_name = ws.title
            
            found_headers = {}
            missing_headers = []
            
            # Search for headers in first N rows
            for header in required_headers:
                found = False
                for row in range(1, min(search_rows + 1, ws.max_row + 1)):
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row, col).value
                        if cell_value and isinstance(cell_value, str):
                            if header.lower() in cell_value.lower():
                                found_headers[header] = {"row": row, "column": col, "value": cell_value}
                                found = True
                                break
                    if found:
                        break
                
                if not found:
                    missing_headers.append(header)
            
            wb.close()
            
            if missing_headers:
                search_area = f"first {search_rows} rows of '{worksheet_name}'"
                raise HeaderNotFoundError(
                    header_name=", ".join(missing_headers),
                    search_area=search_area
                )
            
            return {
                "worksheet_name": worksheet_name,
                "found_headers": found_headers,
                "search_rows": search_rows
            }
            
        except (HeaderNotFoundError, WorksheetNotFoundError):
            raise
        except Exception as e:
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason=f"Failed to validate headers: {str(e)}"
            )
    
    @classmethod 
    def validate_data_sufficient(cls, file_path: Union[str, Path],
                                min_rows: int = 1,
                                worksheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Validate that worksheet has sufficient data
        
        Args:
            file_path: Path to Excel file
            min_rows: Minimum number of data rows required
            worksheet_name: Name of worksheet to check
            
        Returns:
            Dictionary with data statistics
            
        Raises:
            InsufficientDataError: If not enough data
        """
        path = FileValidator.validate_file_format(file_path)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, skipping data validation")
            return {"worksheet_name": worksheet_name or "unknown", "total_rows": 0, "total_columns": 0, "data_rows": 0, "min_required": min_rows}
        
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True)
            
            if worksheet_name:
                if worksheet_name not in wb.sheetnames:
                    raise WorksheetNotFoundError(worksheet_name, wb.sheetnames)
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                worksheet_name = ws.title
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Count non-empty data rows (skip first few rows that might be headers)
            data_rows = 0
            for row in range(4, max_row + 1):  # Start from row 4 (common header location)
                has_data = False
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row, col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                        break
                if has_data:
                    data_rows += 1
            
            wb.close()
            
            if data_rows < min_rows:
                raise InsufficientDataError(
                    data_type="data rows",
                    required_count=min_rows,
                    actual_count=data_rows
                )
            
            return {
                "worksheet_name": worksheet_name,
                "total_rows": max_row,
                "total_columns": max_col,
                "data_rows": data_rows,
                "min_required": min_rows
            }
            
        except (InsufficientDataError, WorksheetNotFoundError):
            raise
        except Exception as e:
            raise FileAccessError(
                file_path=str(path),
                operation="read",
                reason=f"Failed to validate data: {str(e)}"
            )


def validate_step1_template(file_path: Union[str, Path]) -> bool:
    """
    Validate Step1 template structure
    
    Args:
        file_path: Path to Step1 template file
        
    Returns:
        True if valid
        
    Raises:
        ValidationError: If validation fails
    """
    try:
        # Basic file validation
        FileValidator.validate_file_format(file_path)
        
        # Check for required template structure
        expected_headers = ["Combination", "General Type Component", "Sub-Type Component"]
        
        ExcelStructureValidator.validate_headers_exist(
            file_path, 
            expected_headers, 
            search_rows=15
        )
        
        # Check minimum column count (17 columns A-Q)
        if OPENPYXL_AVAILABLE:
            ExcelStructureValidator.validate_columns_exist(
                file_path,
                [openpyxl.utils.get_column_letter(i) for i in range(1, 18)]  # A-Q
            )
        
        logger.info(f"Step1 template validation passed: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Step1 template validation failed: {e}")
        raise ValidationError(f"Step1 template validation failed: {str(e)}")


def validate_step2_input(step1_file: Union[str, Path], source_file: Union[str, Path], graceful: bool = False) -> Tuple[bool, List[str]]:
    """
    Validate Step2 inputs (Step1 template + source data file) with optional graceful mode
    
    Args:
        step1_file: Path to Step1 template
        source_file: Path to source data file
        graceful: If True, return warnings instead of raising exceptions
        
    Returns:
        Tuple of (is_valid, list_of_warnings)
        
    Raises:
        ValidationError: If validation fails and graceful=False
    """
    warnings = []
    
    try:
        # Validate Step1 template
        validate_step1_template(step1_file)
    except Exception as e:
        warning = f"Step1 template validation issue: {str(e)}"
        warnings.append(warning)
        if not graceful:
            logger.error(f"Step2 input validation failed: {e}")
            raise ValidationError(f"Step2 input validation failed: {str(e)}")
        
    try:
        # Validate source file
        FileValidator.validate_file_format(source_file)
    except Exception as e:
        warning = f"Source file validation issue: {str(e)}"
        warnings.append(warning)
        if not graceful:
            logger.error(f"Step2 input validation failed: {e}")
            raise ValidationError(f"Step2 input validation failed: {str(e)}")
        
    try:
        # Check source has data
        ExcelStructureValidator.validate_data_sufficient(source_file, min_rows=1)
    except Exception as e:
        warning = f"Source data sufficiency issue: {str(e)}"
        warnings.append(warning)
        if not graceful:
            logger.error(f"Step2 input validation failed: {e}")
            raise ValidationError(f"Step2 input validation failed: {str(e)}")
    
    if warnings:
        logger.warning(f"Step2 input validation completed with {len(warnings)} warnings")
        for warning in warnings:
            logger.warning(f"  - {warning}")
    else:
        logger.info(f"Step2 input validation passed: {step1_file}, {source_file}")
    
    return len(warnings) == 0, warnings

def validate_step2_input_legacy(step1_file: Union[str, Path], source_file: Union[str, Path]) -> bool:
    """
    Legacy validate Step2 inputs - strict mode for backward compatibility
    
    Args:
        step1_file: Path to Step1 template
        source_file: Path to source data file
        
    Returns:
        True if valid
        
    Raises:
        ValidationError: If validation fails
    """
    is_valid, warnings = validate_step2_input(step1_file, source_file, graceful=False)
    return is_valid


def validate_step3_input(step2_file: Union[str, Path]) -> bool:
    """
    Validate Step3 input (Step2 output file)
    
    Args:
        step2_file: Path to Step2 output file
        
    Returns:
        True if valid
        
    Raises:
        ValidationError: If validation fails
    """
    try:
        # Basic file validation
        FileValidator.validate_file_format(step2_file)
        
        # Check has article data in rows 1-2
        ExcelStructureValidator.validate_data_sufficient(step2_file, min_rows=1)
        
        logger.info(f"Step3 input validation passed: {step2_file}")
        return True
        
    except Exception as e:
        logger.error(f"Step3 input validation failed: {e}")
        raise ValidationError(f"Step3 input validation failed: {str(e)}")


def validate_step4_input(step3_file: Union[str, Path]) -> bool:
    """
    Validate Step4 input (Step3 output file)
    
    Args:
        step3_file: Path to Step3 output file
        
    Returns:
        True if valid
        
    Raises:
        ValidationError: If validation fails
    """
    try:
        # Basic file validation
        FileValidator.validate_file_format(step3_file)
        
        # Check required columns D, E, F exist
        ExcelStructureValidator.validate_columns_exist(step3_file, ['D', 'E', 'F'])
        
        # Check has sufficient data
        ExcelStructureValidator.validate_data_sufficient(step3_file, min_rows=3)
        
        logger.info(f"Step4 input validation passed: {step3_file}")
        return True
        
    except Exception as e:
        logger.error(f"Step4 input validation failed: {e}")
        raise ValidationError(f"Step4 input validation failed: {str(e)}")


def validate_step5_input(step4_file: Union[str, Path]) -> bool:
    """
    Validate Step5 input (Step4 output file)
    
    Args:
        step4_file: Path to Step4 output file
        
    Returns:
        True if valid
        
    Raises:
        ValidationError: If validation fails
    """
    try:
        # Basic file validation
        FileValidator.validate_file_format(step4_file)
        
        # Check required columns for filtering (H for NA, B,C,D,E,F,I,J for SD comparison)
        required_cols = ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J']
        ExcelStructureValidator.validate_columns_exist(step4_file, required_cols)
        
        # Check has sufficient data
        ExcelStructureValidator.validate_data_sufficient(step4_file, min_rows=3)
        
        logger.info(f"Step5 input validation passed: {step4_file}")
        return True
        
    except Exception as e:
        logger.error(f"Step5 input validation failed: {e}")
        raise ValidationError(f"Step5 input validation failed: {str(e)}")