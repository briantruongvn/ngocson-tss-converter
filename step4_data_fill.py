#!/usr/bin/env python3
"""
Step 4: Data Fill
Fills empty cells in columns D, E, F with values from the cell above (vertical inheritance).
"""

import openpyxl
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional
import argparse
import sys
import shutil

from common.validation import validate_step4_input, FileValidator
from common.exceptions import TSConverterError

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataFiller:
    """
    Data Filler for Step 4
    
    Fills empty cells in specified columns with data from the cell above:
    - Target columns: D, E, F
    - Starting from row 4 (after headers and article data)
    - Forward fill: empty cells inherit value from cell above
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Target columns for data filling
        self.target_columns = ['D', 'E', 'F']
        self.start_row = 4  # Start from row 4 (after headers)
    
    def find_last_data_row(self, worksheet) -> int:
        """
        Find the last row that contains data in any column
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Last row number (1-based) with data
        """
        last_row = worksheet.max_row
        
        # Search backwards from max_row to find actual last row with data
        for row in range(last_row, 0, -1):
            has_data = False
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and (not isinstance(cell_value, str) or cell_value.strip()):
                    has_data = True
                    break
            
            if has_data:
                logger.info(f"Found last data row: {row}")
                return row
        
        logger.warning("No data found in worksheet")
        return self.start_row
    
    def fill_column(self, worksheet, column_letter: str, start_row: int, end_row: int) -> int:
        """
        Fill empty cells in a column with data from the cell above
        
        Args:
            worksheet: openpyxl worksheet object
            column_letter: Column letter (D, E, F)
            start_row: Starting row (1-based)
            end_row: Ending row (1-based)
            
        Returns:
            Number of cells filled
        """
        col_num = openpyxl.utils.column_index_from_string(column_letter)
        filled_count = 0
        
        logger.info(f"Filling column {column_letter} from row {start_row} to {end_row}")
        
        # Process each row from start to end
        for row in range(start_row, end_row + 1):
            current_cell = worksheet.cell(row, col_num)
            current_value = current_cell.value
            
            # Check if current cell is empty
            is_empty = (current_value is None or 
                       (isinstance(current_value, str) and current_value.strip() == ""))
            
            if is_empty and row > start_row:
                # Get value from cell above
                above_cell = worksheet.cell(row - 1, col_num)
                above_value = above_cell.value
                
                if above_value is not None:
                    # Fill current cell with value from above
                    current_cell.value = above_value
                    filled_count += 1
                    logger.debug(f"Filled {column_letter}{row} with '{above_value}' from {column_letter}{row-1}")
                else:
                    logger.debug(f"Keeping {column_letter}{row} empty (no value above)")
            elif not is_empty:
                logger.debug(f"Keeping {column_letter}{row} = '{current_value}' (has data)")
            else:
                logger.debug(f"Keeping {column_letter}{row} empty (first row or no data above)")
        
        logger.info(f"Column {column_letter}: filled {filled_count} cells")
        return filled_count
    
    def process_columns(self, worksheet) -> dict:
        """
        Process all target columns (D, E, F) and fill empty cells
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Dictionary with fill counts for each column
        """
        # Find the last row with data
        last_row = self.find_last_data_row(worksheet)
        
        if last_row < self.start_row:
            logger.warning(f"No data to process (last_row: {last_row}, start_row: {self.start_row})")
            return {}
        
        results = {}
        total_filled = 0
        
        # Process each target column
        for column_letter in self.target_columns:
            filled_count = self.fill_column(worksheet, column_letter, self.start_row, last_row)
            results[column_letter] = filled_count
            total_filled += filled_count
        
        logger.info(f"Total cells filled: {total_filled}")
        return results
    
    def process_file(self, step3_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step3 file and fill empty cells in columns D, E, F
        
        Args:
            step3_file: Step3 input file path
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 4: Data Fill")
        
        # Validate input file
        try:
            validate_step4_input(step3_file)
            step3_path = Path(step3_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step3_path.stem.replace(" - Step3", "")
            output_file = self.output_dir / f"{base_name} - Step4.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Input: {step3_path}")
        logger.info(f"Output: {output_file}")
        
        # OPTIMIZATION: Load Step3 file directly instead of copying
        logger.info("Loading Step3 file directly (no file copy)")
        
        # Load Step3 workbook
        wb = openpyxl.load_workbook(str(step3_path))
        ws = wb.active
        
        # Process columns
        fill_results = self.process_columns(ws)
        
        # Log results
        if fill_results:
            logger.info("Fill Summary:")
            for column, count in fill_results.items():
                logger.info(f"  Column {column}: {count} cells filled")
        else:
            logger.info("No cells were filled")
        
        # Save output file
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 4 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        wb.close()
        
        return str(output_file)

def main():
    """Command line interface for data filling"""
    parser = argparse.ArgumentParser(description='Data Filler Step 4 - Fill Empty Cells')
    parser.add_argument('step3_file', help='Step3 input file (*.xlsx)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize filler
    filler = DataFiller(args.base_dir)
    
    try:
        result = filler.process_file(args.step3_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()