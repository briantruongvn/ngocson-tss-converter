#!/usr/bin/env python3
"""
Step 2: Data Extraction from Excel Files
Extracts Article Name and Article Number from input Excel files and populates Step1 template.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple, Dict
import argparse
import sys
import re

from common.validation import validate_step2_input, FileValidator
from common.exceptions import TSConverterError
from common.quality_reporter import get_global_reporter

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataExtractor:
    """
    Data Extractor for Step 2
    
    Extracts Article Name and Article Number from Excel files:
    - Searches all sheets for "Product name"/"Article name" and "Product number"/"Article number"
    - Extracts vertical data until empty cell
    - Populates Step1 template with extracted data
    - Removes duplicate pairs
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Headers to search for (updated to include both article and product variations)
        self.name_headers = ["Article name", "Product name", "article name", "product name"]
        self.number_headers = ["Article number", "Product number", "article number", "product number"]
    
    def is_cell_hidden(self, worksheet, row_num: int, col_num: int) -> bool:
        """
        Check if a cell is in a hidden row or column
        
        Args:
            worksheet: openpyxl worksheet object
            row_num: Row number (1-based)
            col_num: Column number (1-based)
            
        Returns:
            True if cell is hidden, False otherwise
        """
        try:
            # Check if row is hidden
            row_hidden = worksheet.row_dimensions.get(row_num, None)
            if row_hidden and getattr(row_hidden, 'hidden', False):
                return True
            
            # Check if column is hidden
            col_letter = get_column_letter(col_num)
            col_hidden = worksheet.column_dimensions.get(col_letter, None)
            if col_hidden and getattr(col_hidden, 'hidden', False):
                return True
            
            return False
        except Exception as e:
            logger.debug(f"Error checking if cell {row_num},{col_num} is hidden: {e}")
            return False
    
    def find_header_cells(self, worksheet, headers: List[str]) -> List[Tuple[int, int]]:
        """
        Find cells containing specified headers in a worksheet, skipping hidden cells
        
        Args:
            worksheet: openpyxl worksheet object
            headers: List of header strings to search for
            
        Returns:
            List of (row, col) tuples where headers are found
        """
        found_cells = []
        cells_checked = 0
        hidden_cells_skipped = 0
        max_cells = 10000  # Safety limit to prevent infinite search
        
        try:
            # Search through all cells in the worksheet (with limits)
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # Limit to first 100 rows
                for col_num in range(1, min(worksheet.max_column + 1, 50)):  # Limit to first 50 columns
                    cells_checked += 1
                    if cells_checked > max_cells:
                        logger.warning(f"Header search timeout in {worksheet.title}: checked {cells_checked} cells")
                        break
                    
                    # Skip hidden cells
                    if self.is_cell_hidden(worksheet, row_num, col_num):
                        hidden_cells_skipped += 1
                        logger.debug(f"Skipping hidden cell {worksheet.title}!{row_num},{col_num}")
                        continue
                        
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell_value = self.safe_cell_value(cell)
                        if cell_value:
                            for header in headers:
                                if header.lower() in cell_value.lower():
                                    found_cells.append((cell.row, cell.column))
                                    logger.info(f"Found '{header}' at {worksheet.title}!{cell.coordinate}: {cell_value}")
                                    break
                    except Exception as cell_error:
                        logger.debug(f"Error reading cell {worksheet.title}!{row_num},{col_num}: {cell_error}")
                        continue
                        
                if cells_checked > max_cells:
                    break
                    
        except Exception as e:
            logger.error(f"Error searching headers in {worksheet.title}: {e}")
        
        logger.debug(f"Header search in {worksheet.title}: checked {cells_checked} cells, skipped {hidden_cells_skipped} hidden cells, found {len(found_cells)} matches")
        return found_cells
    
    def find_m_textile_sheets(self, workbook) -> List[str]:
        """
        Find all sheets that contain 'M-Textile' or 'M- textile' (case insensitive)
        
        Args:
            workbook: openpyxl workbook object
            
        Returns:
            List of sheet names that match the M-Textile pattern
        """
        m_textile_sheets = []
        pattern_variations = ['m-textile', 'm- textile', 'm-textile', 'm- textile']
        
        for sheet_name in workbook.sheetnames:
            sheet_name_lower = sheet_name.lower()
            for pattern in pattern_variations:
                if pattern in sheet_name_lower:
                    m_textile_sheets.append(sheet_name)
                    logger.info(f"Found M-Textile sheet: {sheet_name}")
                    break
        
        if not m_textile_sheets:
            logger.warning("No M-Textile sheets found in workbook")
        
        return m_textile_sheets

    def find_product_combination_header(self, worksheet) -> Optional[Tuple[int, int]]:
        """
        Find the 'Product combination' or 'Product information' header in a worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Tuple of (row, col) if found, None otherwise
        """
        search_patterns = ["product combination", "product information"]
        
        try:
            # Search through the worksheet for header patterns (case insensitive)
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # Limit to first 100 rows
                for col_num in range(1, min(worksheet.max_column + 1, 50)):  # Limit to first 50 columns
                    # Skip hidden cells
                    if self.is_cell_hidden(worksheet, row_num, col_num):
                        continue
                        
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell_value = self.safe_cell_value(cell)
                        if cell_value:
                            for pattern in search_patterns:
                                if pattern in cell_value.lower():
                                    logger.info(f"Found '{pattern}' at {worksheet.title}!{cell.coordinate}: {cell_value}")
                                    return (cell.row, cell.column)
                    except Exception as cell_error:
                        logger.debug(f"Error reading cell {worksheet.title}!{row_num},{col_num}: {cell_error}")
                        continue
        except Exception as e:
            logger.error(f"Error searching for product headers in {worksheet.title}: {e}")
        
        logger.warning(f"Product header (combination/information) not found in {worksheet.title}")
        return None

    def find_headers_upward_from_position(self, worksheet, start_row: int, headers: List[str]) -> List[Tuple[int, int]]:
        """
        Search upward from a given position to find specified headers
        
        Args:
            worksheet: openpyxl worksheet object
            start_row: Row number to start searching upward from
            headers: List of header strings to search for (case insensitive)
            
        Returns:
            List of (row, col) tuples where headers are found
        """
        found_cells = []
        
        try:
            # Search upward from start_row to row 1
            for row_num in range(start_row, 0, -1):  # Search upward
                for col_num in range(1, min(worksheet.max_column + 1, 50)):  # Limit to first 50 columns
                    # Skip hidden cells
                    if self.is_cell_hidden(worksheet, row_num, col_num):
                        continue
                        
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell_value = self.safe_cell_value(cell)
                        if cell_value:
                            for header in headers:
                                if header.lower() in cell_value.lower():
                                    found_cells.append((cell.row, cell.column))
                                    logger.info(f"Found '{header}' upward at {worksheet.title}!{cell.coordinate}: {cell_value}")
                                    break
                    except Exception as cell_error:
                        logger.debug(f"Error reading cell {worksheet.title}!{row_num},{col_num}: {cell_error}")
                        continue
                        
        except Exception as e:
            logger.error(f"Error searching headers upward in {worksheet.title}: {e}")
        
        return found_cells

    def safe_cell_value(self, cell) -> str:
        """
        Safely extract cell value, handling formula errors and edge cases
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Safe string value or empty string if error
        """
        try:
            if cell.value is None:
                return ""
            
            # Check for Excel formula errors
            if isinstance(cell.value, str):
                formula_errors = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!', '#ERROR!']
                if any(error in str(cell.value) for error in formula_errors):
                    warning_msg = f"Formula error detected in {cell.coordinate}: {cell.value} - using empty value"
                    logger.warning(warning_msg)
                    get_global_reporter().add_warning(
                        'step2', 'formula_errors',
                        f"Excel formula error in cell {cell.coordinate}",
                        f"Error value: {cell.value}"
                    )
                    return ""
            
            # Handle numeric values
            if isinstance(cell.value, (int, float)):
                return str(cell.value)
            
            # Handle datetime values  
            from datetime import datetime
            if isinstance(cell.value, datetime):
                return cell.value.strftime('%Y-%m-%d %H:%M:%S')
            
            # Convert to string and clean
            return str(cell.value).strip()
        
        except Exception as e:
            logger.warning(f"Error reading cell {getattr(cell, 'coordinate', 'unknown')}: {e} - using empty value")
            return ""
    
    def clean_value(self, value: str) -> str:
        """
        Clean individual value by removing trailing punctuation and whitespace
        
        Args:
            value: Raw value string
            
        Returns:
            Cleaned value string
        """
        if not value or not isinstance(value, str):
            return ""
        
        # Remove whitespace
        cleaned = value.strip()
        
        # Remove trailing punctuation (semicolon, comma, etc.)
        while cleaned and cleaned[-1] in ';,':
            cleaned = cleaned[:-1].strip()
        
        return cleaned
    
    def parse_multi_value_cell(self, value: str) -> List[str]:
        """
        Parse cell value that may contain multiple items separated by delimiters
        
        Args:
            value: Cell value string
            
        Returns:
            List of individual cleaned values
        """
        if not value or not isinstance(value, str):
            return []
        
        # Common delimiters: newline, semicolon, comma
        delimiters = ['\n', ';', ',']
        
        # Try each delimiter
        for delimiter in delimiters:
            if delimiter in value:
                # Split by delimiter and clean each part
                parts = [self.clean_value(part) for part in value.split(delimiter)]
                # Filter out empty strings
                parts = [part for part in parts if part]
                if len(parts) > 1:
                    logger.debug(f"Split '{value}' by '{delimiter}' into {len(parts)} parts")
                    return parts
        
        # No delimiter found, return as single cleaned item
        cleaned = self.clean_value(value)
        return [cleaned] if cleaned else []
    
    def extract_data_vertical(self, worksheet, start_row: int, start_col: int) -> List[str]:
        """
        Extract data vertically from worksheet starting from specified position, skipping hidden cells
        Parse multi-value cells into individual items
        
        Args:
            worksheet: openpyxl worksheet object
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
            
        Returns:
            List of extracted individual data values
        """
        data = []
        current_row = start_row + 1  # Start from next row after header
        max_rows = 1000  # Safety limit to prevent infinite loops
        rows_checked = 0
        hidden_rows_skipped = 0
        
        try:
            while rows_checked < max_rows:
                # Skip hidden cells
                if self.is_cell_hidden(worksheet, current_row, start_col):
                    hidden_rows_skipped += 1
                    logger.debug(f"Skipping hidden cell {worksheet.title}!{current_row},{start_col}")
                    current_row += 1
                    rows_checked += 1
                    continue
                
                cell = worksheet.cell(row=current_row, column=start_col)
                
                # Use safe cell reading to handle formula errors
                value = self.safe_cell_value(cell)
                
                # Check if we've reached end of data
                if not value:
                    logger.debug(f"Stopping extraction at {worksheet.title}!{cell.coordinate}: empty cell")
                    break
                
                if value:
                    # Parse multi-value cells
                    try:
                        parsed_values = self.parse_multi_value_cell(value)
                        data.extend(parsed_values)
                        logger.debug(f"Extracted from {worksheet.title}!{cell.coordinate}: {len(parsed_values)} items: {parsed_values}")
                    except Exception as parse_error:
                        logger.warning(f"Error parsing cell {worksheet.title}!{cell.coordinate}: {parse_error}")
                        # Continue with raw value
                        cleaned_value = self.clean_value(value)
                        if cleaned_value:
                            data.append(cleaned_value)
                
                current_row += 1
                rows_checked += 1
                
                # Check if we're going beyond reasonable worksheet bounds
                if current_row > worksheet.max_row + 100:
                    logger.warning(f"Stopping extraction: exceeded max_row + 100 at row {current_row}")
                    break
                    
        except Exception as e:
            logger.error(f"Error during data extraction at {worksheet.title}!{current_row},{start_col}: {e}")
            # Return what we have so far
        
        logger.debug(f"Finished extracting from column {start_col}, found {len(data)} items, skipped {hidden_rows_skipped} hidden rows")
        return data

    def extract_article_numbers_by_position(self, worksheet, name_positions: List[Tuple[int, int]]) -> List[str]:
        """
        Extract article numbers based on position - article numbers are assumed to be 
        immediately to the right of article names (col + 1)
        
        Args:
            worksheet: openpyxl worksheet object
            name_positions: List of (row, col) tuples for article name headers
            
        Returns:
            List of article numbers extracted from positions (row, col + 1)
        """
        all_numbers = []
        
        for name_row, name_col in name_positions:
            # Article number is assumed to be at (name_row, name_col + 1)
            number_col = name_col + 1
            
            try:
                numbers = self.extract_data_vertical(worksheet, name_row, number_col)
                all_numbers.extend(numbers)
                logger.info(f"Extracted {len(numbers)} article numbers from {worksheet.title}!{worksheet.cell(name_row, number_col).coordinate} (position-based: col {name_col} + 1)")
            except Exception as e:
                logger.warning(f"Failed to extract article numbers from position {worksheet.title}!({name_row}, {number_col}): {e}")
                # Try to find number headers as fallback
                logger.info("Attempting fallback to find 'Article number' headers...")
                try:
                    number_cells = self.find_header_cells(worksheet, self.number_headers)
                    for row, col in number_cells:
                        numbers = self.extract_data_vertical(worksheet, row, col)
                        all_numbers.extend(numbers)
                        logger.info(f"Fallback: Extracted {len(numbers)} numbers from {worksheet.title}!{worksheet.cell(row, col).coordinate}")
                except Exception as fallback_error:
                    logger.warning(f"Fallback also failed: {fallback_error}")
        
        return all_numbers

    def populate_template_with_merged_cells(self, worksheet, unique_names: List[str], unique_numbers: List[str]) -> None:
        """
        Populate Step1 template with merged cells starting from column R
        
        Args:
            worksheet: Step1 worksheet object
            unique_names: List of unique article names
            unique_numbers: List of unique article numbers
        """
        from openpyxl.styles import Alignment, PatternFill
        
        # Define light orange fill
        light_orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        
        max_pairs = max(len(unique_names), len(unique_numbers))
        start_col = 18  # Column R = 18
        
        for i in range(max_pairs):
            col = start_col + i  # R, S, T, etc.
            col_letter = get_column_letter(col)
            
            # Article name: merge R1:R9 and apply 90-degree rotation
            if i < len(unique_names):
                name = unique_names[i]
                # Merge range for article name (rows 1-9)
                merge_range = f"{col_letter}1:{col_letter}9"
                worksheet.merge_cells(merge_range)
                
                # Set the value in the first cell of the merged range
                name_cell = worksheet.cell(row=1, column=col, value=name)
                
                # Apply 90-degree rotation, center alignment, and light orange background
                name_cell.alignment = Alignment(
                    horizontal="center", 
                    vertical="center", 
                    text_rotation=90,
                    wrap_text=True
                )
                name_cell.fill = light_orange_fill
                
                logger.debug(f"Set merged {merge_range} = '{name}' with 90° rotation and light orange fill")
            
            # Article number: set in row 10
            if i < len(unique_numbers):
                number = unique_numbers[i]
                number_cell = worksheet.cell(row=10, column=col, value=number)
                
                # Apply center alignment and light orange background for numbers
                number_cell.alignment = Alignment(horizontal="center", vertical="center")
                number_cell.fill = light_orange_fill
                logger.debug(f"Set {number_cell.coordinate} = '{number}' with light orange fill")
    
    def remove_duplicates(self, name_data: List[str], number_data: List[str]) -> Tuple[List[str], List[str]]:
        """
        Remove duplicate pairs from name and number data
        
        Args:
            name_data: List of article names
            number_data: List of article numbers
            
        Returns:
            Tuple of (unique_names, unique_numbers)
        """
        if len(name_data) != len(number_data):
            # Pad shorter list with empty strings
            max_len = max(len(name_data), len(number_data))
            name_data.extend([""] * (max_len - len(name_data)))
            number_data.extend([""] * (max_len - len(number_data)))
        
        seen_pairs = set()
        unique_names = []
        unique_numbers = []
        
        for name, number in zip(name_data, number_data):
            pair = (name, number)
            if pair not in seen_pairs:
                seen_pairs.add(pair)
                unique_names.append(name)
                unique_numbers.append(number)
            else:
                logger.info(f"Removed duplicate pair: ('{name}', '{number}')")
        
        return unique_names, unique_numbers

    def process_m_textile_file(self, step1_file: Union[str, Path], 
                              source_file: Union[str, Path],
                              output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step1 file and extract data from M-Textile sheets using the new logic
        
        Args:
            step1_file: Step1 template file path
            source_file: Source Excel file to extract data from
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 2: M-Textile Data Extraction (New Logic)")
        
        # Validate input files
        try:
            validate_step2_input(step1_file, source_file)
            step1_path = Path(step1_file)
            source_path = Path(source_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step1_path.stem.replace(" - Step1", "")
            output_file = self.output_dir / f"{base_name} - Step2.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Step1 Template: {step1_path}")
        logger.info(f"Source Data: {source_path}")
        logger.info(f"Output: {output_file}")
        
        # Load Step1 template (preserve formatting)
        step1_wb = openpyxl.load_workbook(str(step1_path))
        step1_ws = step1_wb.active
        
        # Load source file for data extraction
        source_wb = openpyxl.load_workbook(str(source_path))
        
        # Find M-Textile sheets
        m_textile_sheets = self.find_m_textile_sheets(source_wb)
        if not m_textile_sheets:
            logger.warning("No M-Textile sheets found - creating empty output")
            step1_wb.save(str(output_file))
            source_wb.close()
            step1_wb.close()
            return str(output_file)
        
        all_names = []
        all_numbers = []
        
        # Process only M-Textile sheets
        for sheet_name in m_textile_sheets:
            logger.info(f"Processing M-Textile sheet: {sheet_name}")
            try:
                worksheet = source_wb[sheet_name]
                
                # Find "Product combination" or "Product information" header first
                product_info_pos = self.find_product_combination_header(worksheet)
                if not product_info_pos:
                    logger.warning(f"No 'Product combination/information' header found in {sheet_name} - skipping")
                    continue
                
                product_info_row = product_info_pos[0]
                logger.info(f"Found product header at row {product_info_row} in {sheet_name}")
                
                # Search upward from product header for article/product headers
                name_cells = self.find_headers_upward_from_position(worksheet, product_info_row, self.name_headers)
                number_cells = self.find_headers_upward_from_position(worksheet, product_info_row, self.number_headers)
                
                # Extract data from found headers
                for row, col in name_cells:
                    try:
                        names = self.extract_data_vertical(worksheet, row, col)
                        all_names.extend(names)
                        logger.info(f"Extracted {len(names)} names from {sheet_name}!{worksheet.cell(row, col).coordinate}")
                    except Exception as e:
                        logger.error(f"Error extracting names from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                        continue
                
                # Use position-based article number extraction (numbers are right of names)
                logger.info("Using position-based article number extraction (numbers are right of article names)")
                all_numbers.extend(self.extract_article_numbers_by_position(worksheet, name_cells))
                
                # Keep fallback to header-based extraction if position-based fails
                if not all_numbers and number_cells:
                    logger.info("Position-based extraction failed, using fallback to header-based extraction")
                    for row, col in number_cells:
                        try:
                            numbers = self.extract_data_vertical(worksheet, row, col)
                            all_numbers.extend(numbers)
                            logger.info(f"Extracted {len(numbers)} numbers from {sheet_name}!{worksheet.cell(row, col).coordinate} (fallback)")
                        except Exception as e:
                            logger.error(f"Error extracting numbers from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                            continue
                        
            except Exception as e:
                logger.error(f"Error processing M-Textile sheet {sheet_name}: {e}")
                continue
        
        # Remove duplicates
        unique_names, unique_numbers = self.remove_duplicates(all_names, all_numbers)
        
        logger.info(f"Found {len(all_names)} total names, {len(unique_names)} unique")
        logger.info(f"Found {len(all_numbers)} total numbers, {len(unique_numbers)} unique")
        logger.info(f"Creating {max(len(unique_names), len(unique_numbers))} article pairs")
        
        # Populate Step1 template with merged cells starting from column R
        if unique_names or unique_numbers:
            self.populate_template_with_merged_cells(step1_ws, unique_names, unique_numbers)
        else:
            logger.warning("No data extracted from M-Textile sheets")
        
        # Save output file
        try:
            step1_wb.save(str(output_file))
            logger.info(f"✅ Step 2 M-Textile completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        source_wb.close()
        step1_wb.close()
        
        return str(output_file)
    
    def process_file_with_fallbacks(self, step1_file: Union[str, Path], 
                                   source_file: Union[str, Path],
                                   output_file: Optional[Union[str, Path]] = None,
                                   allow_missing_headers: bool = True) -> str:
        """
        Process Step1 file and extract data from source file with graceful fallbacks
        
        Args:
            step1_file: Step1 template file path
            source_file: Source Excel file to extract data from
            output_file: Optional output file path (if None, auto-generate)
            allow_missing_headers: If True, continue processing even if headers are missing
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 2: Data Extraction (with graceful fallbacks)")
        
        # Initialize processing warnings list
        processing_warnings = []
        
        # Validate input files with graceful mode if fallbacks are allowed
        try:
            if allow_missing_headers:
                # Use graceful validation that returns warnings instead of exceptions
                from common.validation import validate_step2_input as validate_graceful
                validation_result = validate_graceful(step1_file, source_file, graceful=True)
                if isinstance(validation_result, tuple):
                    is_valid, validation_warnings = validation_result
                    if validation_warnings:
                        for warning in validation_warnings:
                            get_global_reporter().add_warning('step2', 'validation_warning', warning)
                            processing_warnings.append(warning)
                else:
                    # If graceful validation is not available, use regular validation
                    validate_step2_input(step1_file, source_file)
            else:
                validate_step2_input(step1_file, source_file)
            
            step1_path = Path(step1_file)
            source_path = Path(source_file)
        except TSConverterError as e:
            if allow_missing_headers:
                logger.warning(f"Input validation failed but continuing with fallbacks: {e}")
                get_global_reporter().add_warning('step2', 'validation_failed', str(e))
                processing_warnings.append(f"Input validation failed: {e}")
                step1_path = Path(step1_file)
                source_path = Path(source_file)
            else:
                logger.error(f"Input validation failed: {e}")
                raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step1_path.stem.replace(" - Step1", "")
            output_file = self.output_dir / f"{base_name} - Step2.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Step1 Template: {step1_path}")
        logger.info(f"Source Data: {source_path}")
        logger.info(f"Output: {output_file}")
        
        # Load Step1 template (preserve formatting)
        step1_wb = openpyxl.load_workbook(str(step1_path))
        step1_ws = step1_wb.active
        
        # Load source file for data extraction
        source_wb = openpyxl.load_workbook(str(source_path))
        
        all_names = []
        all_numbers = []
        
        # Process each worksheet in source file
        for sheet_name in source_wb.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            try:
                worksheet = source_wb[sheet_name]
                
                # Find article name headers with graceful fallback
                try:
                    name_cells = self.find_header_cells(worksheet, self.name_headers)
                    if not name_cells and allow_missing_headers:
                        warning_msg = f"No name headers found in sheet {sheet_name} - using empty placeholder"
                        logger.warning(warning_msg)
                        processing_warnings.append(f"Missing name headers in sheet '{sheet_name}'")
                        get_global_reporter().add_warning(
                            'step2', 'missing_headers', 
                            f"Missing article name headers in sheet '{sheet_name}'",
                            "Expected headers: " + ", ".join(self.name_headers)
                        )
                        # Add empty placeholder to maintain structure
                        all_names.append("")
                    else:
                        for row, col in name_cells:
                            try:
                                names = self.extract_data_vertical(worksheet, row, col)
                                all_names.extend(names)
                                logger.info(f"Extracted {len(names)} names from {sheet_name}!{worksheet.cell(row, col).coordinate}")
                            except Exception as e:
                                logger.error(f"Error extracting names from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                                if allow_missing_headers:
                                    processing_warnings.append(f"Failed to extract names from {sheet_name}: {str(e)}")
                                    continue
                                else:
                                    raise
                except Exception as e:
                    logger.error(f"Error finding name headers in sheet {sheet_name}: {e}")
                    if allow_missing_headers:
                        processing_warnings.append(f"Header search failed in sheet '{sheet_name}': {str(e)}")
                    else:
                        raise
                
                # Find article number headers with graceful fallback
                try:
                    number_cells = self.find_header_cells(worksheet, self.number_headers)
                    if not number_cells and allow_missing_headers:
                        logger.warning(f"No number headers found in sheet {sheet_name} - using empty placeholder")
                        processing_warnings.append(f"Missing number headers in sheet '{sheet_name}'")
                        # Add empty placeholder to maintain structure
                        all_numbers.append("")
                    else:
                        for row, col in number_cells:
                            try:
                                numbers = self.extract_data_vertical(worksheet, row, col)
                                all_numbers.extend(numbers)
                                logger.info(f"Extracted {len(numbers)} numbers from {sheet_name}!{worksheet.cell(row, col).coordinate}")
                            except Exception as e:
                                logger.error(f"Error extracting numbers from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                                if allow_missing_headers:
                                    processing_warnings.append(f"Failed to extract numbers from {sheet_name}: {str(e)}")
                                    continue
                                else:
                                    raise
                except Exception as e:
                    logger.error(f"Error finding number headers in sheet {sheet_name}: {e}")
                    if allow_missing_headers:
                        processing_warnings.append(f"Header search failed in sheet '{sheet_name}': {str(e)}")
                    else:
                        raise
                    
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                if allow_missing_headers:
                    processing_warnings.append(f"Failed to process sheet '{sheet_name}': {str(e)}")
                    continue
                else:
                    raise
        
        # If no data was extracted and we're allowing fallbacks, create minimal viable output
        if not all_names and not all_numbers and allow_missing_headers:
            logger.warning("No data extracted from any sheet - creating minimal viable output")
            processing_warnings.append("No product data found in source file")
            all_names = [""]
            all_numbers = [""]
        
        # Remove duplicates
        unique_names, unique_numbers = self.remove_duplicates(all_names, all_numbers)
        
        logger.info(f"Found {len(all_names)} total names, {len(unique_names)} unique")
        logger.info(f"Found {len(all_numbers)} total numbers, {len(unique_numbers)} unique")
        logger.info(f"Creating {max(len(unique_names), len(unique_numbers))} article pairs")
        
        if processing_warnings:
            logger.warning(f"Processing completed with {len(processing_warnings)} warnings:")
            for warning in processing_warnings:
                logger.warning(f"  - {warning}")
        
        # Populate Step1 template
        # Each pair (name, number) goes in one column starting from B
        max_pairs = max(len(unique_names), len(unique_numbers))
        
        for i in range(max_pairs):
            col = i + 2  # Column B = 2, C = 3, etc.
            
            # Row 1: Article name
            if i < len(unique_names):
                name = unique_names[i]
                cell = step1_ws.cell(row=1, column=col, value=name)
                logger.debug(f"Set {cell.coordinate} = '{name}'")
            
            # Row 2: Article number  
            if i < len(unique_numbers):
                number = unique_numbers[i]
                cell = step1_ws.cell(row=2, column=col, value=number)
                logger.debug(f"Set {cell.coordinate} = '{number}'")
        
        # Save output file
        try:
            step1_wb.save(str(output_file))
            if processing_warnings:
                logger.info(f"✅ Step 2 completed with warnings: {output_file}")
            else:
                logger.info(f"✅ Step 2 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        source_wb.close()
        step1_wb.close()
        
        return str(output_file)
    
    def process_file(self, step1_file: Union[str, Path], 
                    source_file: Union[str, Path],
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step1 file and extract data from source file
        
        Args:
            step1_file: Step1 template file path
            source_file: Source Excel file to extract data from
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 2: Data Extraction")
        
        # Validate input files
        try:
            validate_step2_input(step1_file, source_file)
            step1_path = Path(step1_file)
            source_path = Path(source_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step1_path.stem.replace(" - Step1", "")
            output_file = self.output_dir / f"{base_name} - Step2.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Step1 Template: {step1_path}")
        logger.info(f"Source Data: {source_path}")
        logger.info(f"Output: {output_file}")
        
        # Load Step1 template (preserve formatting)
        step1_wb = openpyxl.load_workbook(str(step1_path))
        step1_ws = step1_wb.active
        
        # Load source file for data extraction
        source_wb = openpyxl.load_workbook(str(source_path))
        
        all_names = []
        all_numbers = []
        
        # Process each worksheet in source file
        for sheet_name in source_wb.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            try:
                worksheet = source_wb[sheet_name]
                
                # Find article name headers
                try:
                    name_cells = self.find_header_cells(worksheet, self.name_headers)
                    for row, col in name_cells:
                        try:
                            names = self.extract_data_vertical(worksheet, row, col)
                            all_names.extend(names)
                            logger.info(f"Extracted {len(names)} names from {sheet_name}!{worksheet.cell(row, col).coordinate}")
                        except Exception as e:
                            logger.error(f"Error extracting names from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                            continue
                except Exception as e:
                    logger.error(f"Error finding name headers in sheet {sheet_name}: {e}")
                
                # Find article number headers  
                try:
                    number_cells = self.find_header_cells(worksheet, self.number_headers)
                    for row, col in number_cells:
                        try:
                            numbers = self.extract_data_vertical(worksheet, row, col)
                            all_numbers.extend(numbers)
                            logger.info(f"Extracted {len(numbers)} numbers from {sheet_name}!{worksheet.cell(row, col).coordinate}")
                        except Exception as e:
                            logger.error(f"Error extracting numbers from {sheet_name}!{worksheet.cell(row, col).coordinate}: {e}")
                            continue
                except Exception as e:
                    logger.error(f"Error finding number headers in sheet {sheet_name}: {e}")
                    
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                continue
        
        # Remove duplicates
        unique_names, unique_numbers = self.remove_duplicates(all_names, all_numbers)
        
        logger.info(f"Found {len(all_names)} total names, {len(unique_names)} unique")
        logger.info(f"Found {len(all_numbers)} total numbers, {len(unique_numbers)} unique")
        logger.info(f"Creating {max(len(unique_names), len(unique_numbers))} article pairs")
        
        # Populate Step1 template
        # Each pair (name, number) goes in one column starting from B
        max_pairs = max(len(unique_names), len(unique_numbers))
        
        for i in range(max_pairs):
            col = i + 2  # Column B = 2, C = 3, etc.
            
            # Row 1: Article name
            if i < len(unique_names):
                name = unique_names[i]
                cell = step1_ws.cell(row=1, column=col, value=name)
                logger.debug(f"Set {cell.coordinate} = '{name}'")
            
            # Row 2: Article number  
            if i < len(unique_numbers):
                number = unique_numbers[i]
                cell = step1_ws.cell(row=2, column=col, value=number)
                logger.debug(f"Set {cell.coordinate} = '{number}'")
        
        # Save output file
        try:
            step1_wb.save(str(output_file))
            logger.info(f"✅ Step 2 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        source_wb.close()
        step1_wb.close()
        
        return str(output_file)
    
    def extract_from_step1_source(self, step1_file: Union[str, Path],
                                 output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Extract data from the original source file that was used to create Step1
        
        Args:
            step1_file: Step1 template file path  
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        step1_path = Path(step1_file)
        
        # Try to find the original source file
        # Assume it's in input folder with same base name
        base_name = step1_path.stem.replace(" - Step1", "")
        source_candidates = [
            self.base_dir / "input" / f"{base_name}.xlsx",
            step1_path.parent.parent / "input" / f"{base_name}.xlsx",
            self.base_dir / f"{base_name}.xlsx"
        ]
        
        source_file = None
        for candidate in source_candidates:
            if candidate.exists():
                source_file = candidate
                break
        
        if source_file is None:
            raise FileNotFoundError(f"Could not find source file for {step1_path}")
        
        return self.process_file(step1_file, source_file, output_file)

def main():
    """Command line interface for data extraction"""
    parser = argparse.ArgumentParser(description='Data Extractor Step 2 - Extract Article Data')
    parser.add_argument('step1_file', help='Step1 template file (*.xlsx)')
    parser.add_argument('-s', '--source', help='Source file to extract data from (if not auto-detected)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize extractor
    extractor = DataExtractor(args.base_dir)
    
    try:
        if args.source:
            # Extract from specified source file using new M-Textile logic
            result = extractor.process_m_textile_file(args.step1_file, args.source, args.output)
        else:
            # Auto-detect source file
            result = extractor.extract_from_step1_source(args.step1_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()