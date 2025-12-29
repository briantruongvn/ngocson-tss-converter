#!/usr/bin/env python3
"""
Step 6: Article Name Cross-Reference
Cross-references article names in column Q against article headers (R1-Y1) and marks matches with "X".
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import logging
from pathlib import Path
from typing import Union, Optional, List, Dict, Tuple
import argparse
import sys
import re

from common.validation import FileValidator
from common.exceptions import TSConverterError

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ArticleCrossReference:
    """
    Article Cross-Reference for Step 6
    
    Cross-references article names in column Q against article headers:
    - Parses article lists from column Q (starting Q11)
    - Extracts individual article names from numbered/delimited lists
    - Matches against article headers in row 1 (columns R onwards)
    - Marks matching columns with "X" in the corresponding data row
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Configuration
        self.start_row = 11  # Start processing from row 11
        self.article_list_column = 'Q'  # Column containing article lists
        self.header_row = 1  # Row containing article name headers
        self.article_header_start_col = 'R'  # First column with article headers
        self.match_marker = "X"  # Value to mark matches
    
    def safe_cell_value(self, cell) -> str:
        """
        Safely extract cell value as string
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Safe string value or empty string if error
        """
        try:
            if cell.value is None:
                return ""
            return str(cell.value).strip()
        except Exception as e:
            logger.debug(f"Error reading cell {getattr(cell, 'coordinate', 'unknown')}: {e}")
            return ""
    
    def normalize_article_name(self, name: str) -> str:
        """
        Normalize article name for comparison
        
        Args:
            name: Raw article name
            
        Returns:
            Normalized article name
        """
        if not name:
            return ""
        
        # Convert to lowercase and strip whitespace
        normalized = name.lower().strip()
        
        # Remove extra whitespace
        normalized = re.sub(r'\s+', ' ', normalized)
        
        return normalized
    
    def parse_article_list(self, cell_value: str) -> List[str]:
        """
        Parse article list from cell value
        
        Handles formats like:
        - "1.STUK stor case 34x51x28 white/black;"
        - "2. STUK stor case 34x51x28 white/black AP;"
        - Multi-line with numbers or semicolons
        
        Args:
            cell_value: Raw cell value containing article list
            
        Returns:
            List of individual article names
        """
        if not cell_value or not isinstance(cell_value, str):
            return []
        
        articles = []
        
        # First, split by newlines and semicolons
        delimiters = ['\n', ';']
        lines = [cell_value]
        
        for delimiter in delimiters:
            new_lines = []
            for line in lines:
                new_lines.extend(line.split(delimiter))
            lines = new_lines
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Remove numbering pattern (1., 2., etc.) from start
            # Pattern: optional whitespace, number, dot, optional space
            numbered_pattern = r'^\s*\d+\.\s*'
            clean_line = re.sub(numbered_pattern, '', line)
            
            # Remove trailing punctuation (semicolons, etc.)
            clean_line = clean_line.rstrip(';,').strip()
            
            if clean_line:
                articles.append(clean_line)
                logger.debug(f"Parsed article: '{clean_line}' from '{line}'")
        
        logger.debug(f"Parsed {len(articles)} articles from cell value")
        return articles
    
    def find_article_headers(self, worksheet) -> Dict[str, int]:
        """
        Find article name headers in row 1 starting from column R
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Dictionary mapping normalized article names to column numbers
        """
        article_headers = {}
        start_col = column_index_from_string(self.article_header_start_col)
        
        # Check columns R onwards until we find no more headers
        col = start_col
        empty_count = 0
        max_empty = 5  # Stop after 5 consecutive empty columns
        
        while empty_count < max_empty and col <= worksheet.max_column + 10:
            try:
                cell = worksheet.cell(row=self.header_row, column=col)
                header_value = self.safe_cell_value(cell)
                
                if header_value:
                    normalized_name = self.normalize_article_name(header_value)
                    if normalized_name:
                        article_headers[normalized_name] = col
                        logger.debug(f"Found article header: '{header_value}' -> '{normalized_name}' at column {get_column_letter(col)}")
                        empty_count = 0  # Reset empty counter
                    else:
                        empty_count += 1
                else:
                    empty_count += 1
                
            except Exception as e:
                logger.debug(f"Error reading header at column {col}: {e}")
                empty_count += 1
            
            col += 1
        
        logger.info(f"Found {len(article_headers)} article headers in row {self.header_row}")
        return article_headers
    
    def find_matches(self, article_name: str, article_headers: Dict[str, int]) -> List[int]:
        """
        Find matching columns for an article name
        
        Args:
            article_name: Article name to find matches for
            article_headers: Dictionary of normalized header names to column numbers
            
        Returns:
            List of column numbers that match
        """
        matches = []
        normalized_name = self.normalize_article_name(article_name)
        
        if not normalized_name:
            return matches
        
        # First try exact match
        if normalized_name in article_headers:
            matches.append(article_headers[normalized_name])
            logger.debug(f"Exact match found for '{article_name}' -> column {get_column_letter(article_headers[normalized_name])}")
            return matches
        
        # Try partial matching (contains)
        for header_name, col_num in article_headers.items():
            if normalized_name in header_name or header_name in normalized_name:
                matches.append(col_num)
                logger.debug(f"Partial match found for '{article_name}' with header '{header_name}' -> column {get_column_letter(col_num)}")
        
        if not matches:
            logger.debug(f"No matches found for article: '{article_name}'")
        
        return matches
    
    def mark_matches(self, worksheet, row_num: int, matching_columns: List[int]) -> int:
        """
        Mark matching columns with "X" in the specified row
        
        Args:
            worksheet: openpyxl worksheet object
            row_num: Row number to mark matches in
            matching_columns: List of column numbers to mark
            
        Returns:
            Number of cells marked
        """
        marked_count = 0
        
        for col_num in matching_columns:
            try:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = self.match_marker
                marked_count += 1
                logger.debug(f"Marked {cell.coordinate} with '{self.match_marker}'")
            except Exception as e:
                logger.warning(f"Error marking cell at row {row_num}, column {col_num}: {e}")
        
        return marked_count
    
    def clear_article_lists(self, worksheet) -> int:
        """
        Clear article list data from column Q starting from start_row
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Number of cells cleared
        """
        logger.info("Step 6 Sub-step: Clearing article names from column Q")
        
        article_list_col = column_index_from_string(self.article_list_column)
        cleared_count = 0
        
        current_row = self.start_row
        
        while current_row <= worksheet.max_row:
            try:
                # Get article list cell
                list_cell = worksheet.cell(row=current_row, column=article_list_col)
                
                # Check if cell has content
                if list_cell.value is not None:
                    # Clear the cell
                    list_cell.value = None
                    cleared_count += 1
                    logger.debug(f"Cleared article list from Q{current_row}")
                
            except Exception as e:
                logger.warning(f"Error clearing cell Q{current_row}: {e}")
            
            current_row += 1
        
        logger.info(f"Cleared {cleared_count} article list cells from column Q")
        return cleared_count
    
    def process_file(self, step5_file: Union[str, Path], 
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step 5 file and add article cross-references
        
        Args:
            step5_file: Step 5 input file path
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to output file
        """
        logger.info("📋 Step 6: Article Name Cross-Reference")
        
        # Validate input file
        try:
            step5_path = Path(step5_file)
            if not step5_path.exists():
                raise FileNotFoundError(f"Step 5 file not found: {step5_path}")
            
            # Basic file validation
            FileValidator.validate_file_format(step5_path)
            
        except Exception as e:
            logger.error(f"Input validation failed: {e}")
            raise TSConverterError(f"Invalid Step 5 file: {e}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = step5_path.stem.replace(" - Step5", "")
            output_file = self.output_dir / f"Standard Internal TSS - {base_name}.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except Exception as e:
            logger.error(f"Output validation failed: {e}")
            raise TSConverterError(f"Invalid output path: {e}")
        
        logger.info(f"Input: {step5_path}")
        logger.info(f"Output: {output_file}")
        
        # Load Step 5 file
        try:
            workbook = openpyxl.load_workbook(str(step5_path))
            worksheet = workbook.active
        except Exception as e:
            logger.error(f"Failed to load Step 5 file: {e}")
            raise TSConverterError(f"Cannot load Excel file: {e}")
        
        # Find article headers in row 1
        article_headers = self.find_article_headers(worksheet)
        if not article_headers:
            logger.warning("No article headers found - output will have no cross-references")
        
        # Process article lists starting from specified row
        article_list_col = column_index_from_string(self.article_list_column)
        total_matches = 0
        processed_rows = 0
        
        current_row = self.start_row
        
        while current_row <= worksheet.max_row:
            try:
                # Get article list from column Q
                list_cell = worksheet.cell(row=current_row, column=article_list_col)
                list_value = self.safe_cell_value(list_cell)
                
                if not list_value:
                    # Skip empty rows but continue checking
                    current_row += 1
                    continue
                
                logger.debug(f"Processing row {current_row}: '{list_value[:50]}...'")
                
                # Parse article list from cell
                articles = self.parse_article_list(list_value)
                
                if not articles:
                    logger.debug(f"No articles parsed from row {current_row}")
                    current_row += 1
                    continue
                
                # Find matches for each article
                all_matching_columns = []
                for article in articles:
                    matching_cols = self.find_matches(article, article_headers)
                    all_matching_columns.extend(matching_cols)
                
                # Remove duplicates while preserving order
                unique_matching_columns = []
                seen = set()
                for col in all_matching_columns:
                    if col not in seen:
                        unique_matching_columns.append(col)
                        seen.add(col)
                
                # Mark matches in current row
                if unique_matching_columns:
                    marked = self.mark_matches(worksheet, current_row, unique_matching_columns)
                    total_matches += marked
                    logger.info(f"Row {current_row}: Found {len(articles)} articles, marked {marked} columns")
                else:
                    logger.debug(f"Row {current_row}: Found {len(articles)} articles but no matches")
                
                processed_rows += 1
                
            except Exception as e:
                logger.error(f"Error processing row {current_row}: {e}")
                # Continue with next row
            
            current_row += 1
        
        logger.info(f"Processed {processed_rows} rows with article lists")
        logger.info(f"Total matches marked: {total_matches}")
        
        # Sub-step: Clear article names from column Q
        cleared_count = self.clear_article_lists(worksheet)
        logger.info(f"Sub-step completed: Cleared {cleared_count} article list cells")
        
        # Save output file
        try:
            workbook.save(str(output_file))
            logger.info(f"✅ Step 6 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise TSConverterError(f"Cannot save output file: {e}")
        finally:
            workbook.close()
        
        return str(output_file)

def main():
    """Command line interface for article cross-reference"""
    parser = argparse.ArgumentParser(description='Article Cross-Reference Step 6 - Mark Article Matches')
    parser.add_argument('step5_file', help='Step 5 file (*.xlsx)')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize cross-reference processor
    crossref = ArticleCrossReference(args.base_dir)
    
    try:
        result = crossref.process_file(args.step5_file, args.output)
        
        print(f"\n✅ Success!")
        print(f"📁 Output: {result}")
        
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()