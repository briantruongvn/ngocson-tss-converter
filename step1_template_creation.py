#!/usr/bin/env python3
"""
Step 1: Create Initial Template
Creates structured output template for Excel format conversion.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional
import argparse
import sys
import re

from common.validation import FileValidator, validate_step1_template
from common.exceptions import TSConverterError
from common.config import get_config

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TemplateCreator:
    """
    Standalone Template Creator for Step 1
    
    Creates structured output template with predefined headers:
    - Row 1: Article name
    - Row 2: Article number  
    - Row 3: 17 column headers (A-Q)
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.config = get_config()
        
        # Set up directories from config
        if base_dir:
            self.base_dir = Path(base_dir)
        else:
            self.base_dir = Path(self.config.get("general.base_dir", "."))
        
        self.output_dir = self.base_dir / self.config.get("general.output_dir", "output")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Template structure from configuration
        self.template_headers = self.config.get("step1.template_headers", [])
        
        # Define styles
        self.row1_2_style = {
            "font": Font(bold=True, color="00000000"),
            "fill": PatternFill(start_color="00B8E6B8", end_color="00B8E6B8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        }
        
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def create_template(self, input_file: Union[str, Path], 
                       output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Create output template from input Excel file
        
        Args:
            input_file: Input Excel file (.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to template file
        """
        logger.info("📋 Step 1: Create Initial Template")
        
        # Validate input file format
        try:
            input_path = FileValidator.validate_file_format(input_file)
        except TSConverterError as e:
            logger.error(f"Input validation failed: {e}")
            raise
        
        # Auto-generate output file if not provided
        if output_file is None:
            base_name = input_path.stem
            output_file = self.output_dir / f"{base_name} - Step1.xlsx"
        else:
            output_file = Path(output_file)
        
        # Validate output path is writable
        try:
            output_file = FileValidator.validate_output_writable(output_file)
        except TSConverterError as e:
            logger.error(f"Output validation failed: {e}")
            raise
        
        logger.info(f"Input: {input_path}")
        logger.info(f"Output: {output_file}")
        
        # Create new workbook with template structure
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Output Template"
        
        # Row 1: Article name with formatting
        cell1 = ws.cell(1, 1, "Article name")
        cell1.font = self.row1_2_style["font"]
        cell1.fill = self.row1_2_style["fill"]
        cell1.alignment = self.row1_2_style["alignment"]
        
        # Row 2: Article number with formatting
        cell2 = ws.cell(2, 1, "Article number")
        cell2.font = self.row1_2_style["font"]
        cell2.fill = self.row1_2_style["fill"]
        cell2.alignment = self.row1_2_style["alignment"]
        
        # Row 3: Headers (17 columns A-Q) with specific formatting and column widths
        for col_idx, header_info in enumerate(self.template_headers, 1):
            cell = ws.cell(3, col_idx, header_info["name"])
            
            # Apply font with specific color for each column
            cell.font = Font(bold=True, color=header_info["font_color"])
            
            # Apply background color
            cell.fill = PatternFill(start_color=header_info["bg_color"], 
                                   end_color=header_info["bg_color"], 
                                   fill_type="solid")
            
            # Apply alignment
            cell.alignment = self.header_alignment
            
            # Set column width
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = header_info["width"]
        
        logger.info(f"✅ Created formatted template with {len(self.template_headers)} headers")
        
        # Save template
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 1 completed: {output_file}")
            
            # Validate created template structure
            validate_step1_template(output_file)
            logger.info("✅ Template validation passed")
            
        except Exception as e:
            logger.error(f"Failed to save or validate file: {e}")
            raise
        
        return str(output_file)
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step2.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def create_multiple_templates(self, input_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Create templates for multiple files matching patterns
        
        Args:
            input_patterns: List of file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in input_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                input_files = list(self.base_dir.glob(str(pattern)))
            else:
                input_files = [Path(pattern)]
            
            for input_file in input_files:
                if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.create_template(input_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {input_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {input_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {input_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone template creation"""
    parser = argparse.ArgumentParser(description='Template Creator Step 1 - Initial Template Creation')
    parser.add_argument('input', nargs='+', help='Input Excel file(s) (.xlsx format)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize creator
    creator = TemplateCreator(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1:
            # Multiple files mode
            output_dir = args.output if args.output else None
            results = creator.create_multiple_templates(args.input, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            input_file = args.input[0]
            output_file = args.output
            
            result = creator.create_template(input_file, output_file)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()