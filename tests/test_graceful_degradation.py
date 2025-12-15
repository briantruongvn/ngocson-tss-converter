"""
Test Graceful Degradation Features
Tests the new "no crash" strategy for missing headers and formula errors
"""

import unittest
import tempfile
import openpyxl
from pathlib import Path
import sys
import os

# Add parent directory to Python path
sys.path.insert(0, str(Path(__file__).parent.parent))

from step2_data_extraction import DataExtractor
from common.quality_reporter import get_global_reporter, reset_global_reporter


class TestGracefulDegradation(unittest.TestCase):
    """Test graceful degradation features"""
    
    def setUp(self):
        """Set up test environment"""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.extractor = DataExtractor()
        reset_global_reporter()
        
    def tearDown(self):
        """Clean up test environment"""
        import shutil
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel_with_formulas(self, filename: str, include_formula_errors: bool = True):
        """Create test Excel file with formula errors"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add headers (correctly positioned)
        ws['A1'] = "Product name"
        ws['B1'] = "Article number"
        
        # Add some good data
        ws['A2'] = "Product A"
        ws['B2'] = "PRD-001"
        
        # Add formula errors if requested
        if include_formula_errors:
            ws['A3'] = "#N/A"  # Formula error
            ws['B3'] = "PRD-002"
            
            ws['A4'] = "#REF!"  # Another formula error
            ws['B4'] = "PRD-003"
            
            ws['A5'] = "#VALUE!"  # Yet another formula error
            ws['B5'] = "PRD-004"
        
        file_path = self.temp_dir / filename
        wb.save(str(file_path))
        wb.close()
        return file_path
    
    def create_test_excel_missing_headers(self, filename: str):
        """Create test Excel file with missing headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some data but NO headers
        ws['A1'] = "Product A"
        ws['B1'] = "PRD-001"
        ws['A2'] = "Product B"  
        ws['B2'] = "PRD-002"
        
        file_path = self.temp_dir / filename
        wb.save(str(file_path))
        wb.close()
        return file_path
        
    def create_step1_template(self, filename: str):
        """Create a simple Step1 template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create minimal template structure (17 columns A-Q)
        headers = [
            "Combination", "General Type Component(Type)", "Sub-Type Component Identity Process Name",
            "Material Designation", "Material Distributor", "Producer", "Material Type In Process",
            "Document type", "Requirement Source/TED", "Sub-type", "Regulation or substances",
            "Limit", "Test method", "Frequency", "Level", "Warning Limit", "Additional Information"
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=header)
        
        file_path = self.temp_dir / filename
        wb.save(str(file_path))
        wb.close()
        return file_path
    
    def test_safe_cell_value_with_formula_errors(self):
        """Test that safe_cell_value handles formula errors gracefully"""
        # Create Excel file with formula errors
        excel_file = self.create_test_excel_with_formulas("test_formula_errors.xlsx")
        
        # Load workbook and test safe cell reading
        wb = openpyxl.load_workbook(str(excel_file))
        ws = wb.active
        
        # Test normal cell
        normal_cell = ws['A2']  # "Product A"
        result = self.extractor.safe_cell_value(normal_cell)
        self.assertEqual(result, "Product A")
        
        # Test formula error cells
        error_cell1 = ws['A3']  # "#N/A"
        result1 = self.extractor.safe_cell_value(error_cell1)
        self.assertEqual(result1, "")  # Should return empty string
        
        error_cell2 = ws['A4']  # "#REF!"
        result2 = self.extractor.safe_cell_value(error_cell2)
        self.assertEqual(result2, "")  # Should return empty string
        
        error_cell3 = ws['A5']  # "#VALUE!"
        result3 = self.extractor.safe_cell_value(error_cell3)
        self.assertEqual(result3, "")  # Should return empty string
        
        wb.close()
        
        # Check that warnings were logged to quality reporter
        reporter = get_global_reporter()
        warnings = reporter.get_issues_by_category('formula_errors')
        self.assertGreater(len(warnings), 0)  # Should have logged formula error warnings
        
    def test_graceful_missing_headers(self):
        """Test graceful handling of missing headers"""
        # Create files
        step1_file = self.create_step1_template("step1_template.xlsx")
        source_file = self.create_test_excel_missing_headers("missing_headers.xlsx")
        
        # Test graceful processing - should not crash
        try:
            output_file = self.extractor.process_file_with_fallbacks(
                str(step1_file),
                str(source_file),
                allow_missing_headers=True
            )
            
            # Should complete without exception
            self.assertTrue(Path(output_file).exists())
            
            # Check that warnings were logged
            reporter = get_global_reporter()
            warnings = reporter.get_issues_by_category('missing_headers')
            self.assertGreater(len(warnings), 0)  # Should have logged missing header warnings
            
        except Exception as e:
            self.fail(f"Graceful processing should not raise exceptions: {e}")
    
    def test_formula_errors_processing_continues(self):
        """Test that processing continues when formula errors are encountered"""
        # Create files
        step1_file = self.create_step1_template("step1_template.xlsx")
        source_file = self.create_test_excel_with_formulas("formula_errors.xlsx", include_formula_errors=True)
        
        # Test graceful processing 
        try:
            output_file = self.extractor.process_file_with_fallbacks(
                str(step1_file),
                str(source_file),
                allow_missing_headers=True
            )
            
            # Should complete without exception
            self.assertTrue(Path(output_file).exists())
            
            # Verify output file has some data (non-error values)
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            
            # Should have "Product A" in row 1 (first valid entry)
            found_product_a = False
            for col in range(1, 10):  # Check first 10 columns
                if ws.cell(row=1, column=col).value == "Product A":
                    found_product_a = True
                    break
                    
            self.assertTrue(found_product_a, "Should have extracted valid product data")
            
            wb.close()
            
            # Check that formula error warnings were logged
            reporter = get_global_reporter()
            formula_warnings = reporter.get_issues_by_category('formula_errors')
            self.assertGreater(len(formula_warnings), 0)  # Should have logged formula error warnings
            
        except Exception as e:
            self.fail(f"Processing with formula errors should not crash: {e}")
    
    def test_quality_reporter_tracks_issues(self):
        """Test that quality reporter correctly tracks and reports issues"""
        reset_global_reporter()
        reporter = get_global_reporter()
        
        # Start processing
        reporter.start_processing()
        
        # Add some test issues
        reporter.add_warning('step2', 'missing_headers', 'Test missing header warning')
        reporter.add_warning('step2', 'formula_errors', 'Test formula error warning')
        reporter.add_error('step3', 'processing_failed', 'Test processing error')
        
        # End processing
        reporter.end_processing()
        
        # Check summary
        summary = reporter.get_user_summary()
        
        self.assertEqual(summary['warnings_count'], 2)
        self.assertEqual(summary['errors_count'], 1)
        self.assertEqual(summary['total_issues'], 3)
        self.assertLess(summary['quality_score'], 100)  # Score should be reduced due to issues
        
        # Check issue categorization
        warnings = reporter.get_issues_by_level('warning')
        errors = reporter.get_issues_by_level('error')
        
        self.assertEqual(len(warnings), 2)
        self.assertEqual(len(errors), 1)
        
    def test_minimum_viable_output_creation(self):
        """Test that system creates minimal viable output even with major issues"""
        # Create completely empty Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = ""  # Completely empty
        
        empty_file = self.temp_dir / "empty.xlsx"
        wb.save(str(empty_file))
        wb.close()
        
        step1_file = self.create_step1_template("step1_template.xlsx")
        
        # Should create output even with empty input
        try:
            output_file = self.extractor.process_file_with_fallbacks(
                str(step1_file),
                str(empty_file),
                allow_missing_headers=True
            )
            
            # Should complete and create a file
            self.assertTrue(Path(output_file).exists())
            
            # Check quality score reflects the poor input quality
            reporter = get_global_reporter()
            summary = reporter.get_user_summary()
            self.assertLess(summary['quality_score'], 80)  # Should be reduced score due to multiple issues
            self.assertGreater(summary['warnings_count'], 0)  # Should have warnings
            
        except Exception as e:
            self.fail(f"Should create minimal viable output even with empty input: {e}")

if __name__ == '__main__':
    unittest.main()